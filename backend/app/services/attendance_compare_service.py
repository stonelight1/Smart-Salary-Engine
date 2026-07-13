"""
考勤人员比对服务

负责：
- 导入考勤人员 Excel
- 自动比对现有员工档案
- 生成员工异动待确认记录（入职/离职/调岗/冲突）
"""

import json
import logging
import uuid
from datetime import date, datetime
from decimal import Decimal

from app.core.exceptions import BizError
from app.db.database import SessionLocal
from app.models import (
    EmployeeMaster, EmployeeChangeCandidate, EmployeeAssignmentHistory,
    ImportBatch, WorkbookFile,
)
from app.services.employee_service import _unique_employee_no
from app.engines.importer.importer import validate_file, save_file, parse_workbook

logger = logging.getLogger(__name__)


def import_attendance_employees(
    file_bytes: bytes,
    original_name: str,
    salary_month: str,
    created_by: str,
) -> dict:
    """
    导入考勤人员 Excel，比对员工档案，生成异动记录。

    Returns:
        {
            "batch_id": "...",
            "matched_count": int,
            "new_hire_count": int,
            "possible_terminations": int,
            "dept_changes": int,
            "pos_changes": int,
            "conflicts": int,
            "summary_count": int,
            "empty_count": int,
            "candidates": [{...}],
            "unmatched_names": [str],
        }
    """
    validate_file(file_bytes, original_name)
    storage_path, file_hash = save_file(file_bytes, original_name, "attendance")
    parse_result = parse_workbook(storage_path)

    if parse_result["sheet_count"] == 0:
        raise BizError(error_code="EXCEL_PARSE_FAILED", message="未找到工作表")

    sheet = parse_result["sheets"][0]
    headers = sheet["headers"]
    header_row = sheet["header_row_index"]

    if not headers:
        raise BizError(error_code="EXCEL_PARSE_FAILED", message="未识别到表头")

    # 自动检测字段列
    name_col = _detect_col(headers, ["姓名", "员工姓名", "名字", "人员姓名"])
    emp_no_col = _detect_col(headers, ["员工编号", "工号", "编号", "员工ID"])
    dept_col = _detect_col(headers, ["部门", "所属部门", "组织"])
    pos_col = _detect_col(headers, ["岗位", "职位", "职务"])
    hire_col = _detect_col(headers, ["入职日期", "入职时间", "入职"])

    if name_col is None:
        raise BizError(error_code="EXCEL_PARSE_FAILED", message="未识别到员工姓名字段")

    db = SessionLocal()
    batch_id = f"batch_{uuid.uuid4().hex[:8]}"
    try:
        from openpyxl import load_workbook
        wb = load_workbook(storage_path, data_only=True)
        ws = wb[sheet["sheet_name"]]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        data_start = (header_row or 0) + 1
        candidates = []
        matched_count = 0
        new_hire_count = 0
        summary_count = 0
        empty_count = 0
        seen_emp_nos = {}

        for idx in range(data_start, len(all_rows)):
            row = all_rows[idx]
            if not row:
                empty_count += 1
                continue

            # 空行检查
            if all(v is None for v in row):
                empty_count += 1
                continue

            name_val = str(row[name_col]).strip() if name_col < len(row) and row[name_col] is not None else ""

            # 汇总关键词
            if any(kw in name_val for kw in ["合计", "总计", "汇总", "小计", "总结", "自动总结", "总汇总"]):
                summary_count += 1
                continue

            # 无效姓名过滤：空值、None、null、undefined、纯空格
            if not name_val or name_val.strip() == "":
                empty_count += 1
                continue

            # 排除字符串形式的 null/None/undefined
            if name_val.strip().lower() in ("none", "null", "undefined"):
                empty_count += 1
                continue

            employee_no = str(row[emp_no_col]).strip() if emp_no_col is not None and emp_no_col < len(row) and row[emp_no_col] is not None else ""
            department = str(row[dept_col]).strip() if dept_col is not None and dept_col < len(row) and row[dept_col] is not None else ""
            position = str(row[pos_col]).strip() if pos_col is not None and pos_col < len(row) and row[pos_col] is not None else ""
            hire_date_raw = row[hire_col] if hire_col is not None and hire_col < len(row) else None

            # 处理入职日期
            hire_date = None
            if hire_date_raw:
                try:
                    if isinstance(hire_date_raw, (int, float)):
                        from datetime import timedelta
                        base = datetime(1899, 12, 30)
                        hire_date = (base + timedelta(days=int(hire_date_raw))).date()
                    else:
                        hire_date = datetime.strptime(str(hire_date_raw)[:10], "%Y-%m-%d").date()
                except Exception:
                    pass

            # 尝试匹配现有员工
            emp = _find_employee(db, employee_no, name_val)

            if emp:
                # 已匹配 —— 检查变化
                changes = []

                # 部门变化
                if department and emp.department != department:
                    changes.append({
                        "type": "DEPARTMENT_CHANGE",
                        "old": emp.department or "",
                        "new": department,
                    })

                # 岗位变化
                if position and emp.position != position:
                    changes.append({
                        "type": "POSITION_CHANGE",
                        "old": emp.position or "",
                        "new": position,
                    })

                if changes:
                    cand = EmployeeChangeCandidate(
                        id=f"cand_{uuid.uuid4().hex[:8]}",
                        salary_month=salary_month,
                        employee_master_id=emp.id,
                        candidate_type="DEPARTMENT_CHANGE" if any(c["type"] == "DEPARTMENT_CHANGE" for c in changes) else "POSITION_CHANGE",
                        employee_no=employee_no or emp.employee_no or "",
                        employee_name=name_val,
                        department=department or emp.department or "",
                        position=position or emp.position or "",
                        hire_date=hire_date or emp.hire_date,
                        old_data_json=json.dumps({"department": emp.department, "position": emp.position}, ensure_ascii=False),
                        new_data_json=json.dumps({"department": department, "position": position}, ensure_ascii=False),
                        detection_reason="; ".join(f"{c['type']}: {c['old']}→{c['new']}" for c in changes),
                        source_batch_id=batch_id,
                        status="PENDING",
                    )
                    db.add(cand)
                    candidates.append({
                        "type": cand.candidate_type,
                        "employee_name": name_val,
                        "old": emp.department or emp.position or "",
                        "new": department or position or "",
                    })

                # 更新最后出现月份
                emp.last_seen_month = salary_month
                matched_count += 1

            else:
                # 未匹配 —— 疑似新入职
                # 检查同姓名冲突
                dup = db.query(EmployeeMaster).filter(
                    EmployeeMaster.employee_name == name_val,
                ).count()

                cand_type = "POSSIBLE_HIRE"
                conflict_reason = ""
                if employee_no:
                    no_dup = db.query(EmployeeMaster).filter(
                        EmployeeMaster.employee_no == employee_no,
                        EmployeeMaster.id != None,
                    ).count()
                    if no_dup > 0:
                        cand_type = "INFO_CONFLICT"
                        conflict_reason = f"员工编号{employee_no}已存在但姓名不匹配"

                if dup > 0 and cand_type != "INFO_CONFLICT":
                    cand_type = "INFO_CONFLICT"
                    conflict_reason = f"同名员工{name_val}存在{dup}人，无法自动匹配"

                cand = EmployeeChangeCandidate(
                    id=f"cand_{uuid.uuid4().hex[:8]}",
                    salary_month=salary_month,
                    candidate_type=cand_type,
                    employee_no=employee_no,
                    employee_name=name_val,
                    department=department,
                    position=position,
                    hire_date=hire_date,
                    old_data_json="{}",
                    new_data_json=json.dumps({
                        "employee_no": employee_no,
                        "employee_name": name_val,
                        "department": department,
                        "position": position,
                        "hire_date": str(hire_date) if hire_date else "",
                    }, ensure_ascii=False),
                    detection_reason=conflict_reason or f"考表发现新员工: {name_val}",
                    source_batch_id=batch_id,
                    status="PENDING",
                )
                db.add(cand)

                if cand_type == "POSSIBLE_HIRE":
                    new_hire_count += 1
                candidates.append({
                    "type": cand_type,
                    "employee_name": name_val,
                })

        # 检测疑似离职：在职员工不在本月考勤表中
        active_emps = db.query(EmployeeMaster).filter(
            EmployeeMaster.status == "ACTIVE",
        ).all()

        attendance_names = set()
        attendance_emp_nos = set()
        for idx in range(data_start, len(all_rows)):
            row = all_rows[idx]
            if not row:
                continue
            name_v = str(row[name_col]).strip() if name_col < len(row) and row[name_col] is not None else ""
            if name_v:
                attendance_names.add(name_v)
            if emp_no_col is not None:
                en = str(row[emp_no_col]).strip() if emp_no_col < len(row) and row[emp_no_col] is not None else ""
                if en:
                    attendance_emp_nos.add(en)

        possible_terminations = 0
        for emp in active_emps:
            if emp.employee_no and emp.employee_no in attendance_emp_nos:
                continue
            if emp.employee_name in attendance_names:
                continue
            # 不在考勤表中
            cand = EmployeeChangeCandidate(
                id=f"cand_{uuid.uuid4().hex[:8]}",
                salary_month=salary_month,
                employee_master_id=emp.id,
                candidate_type="POSSIBLE_TERMINATION",
                employee_no=emp.employee_no or "",
                employee_name=emp.employee_name,
                department=emp.department or "",
                position=emp.position or "",
                old_data_json=json.dumps({"status": "ACTIVE"}, ensure_ascii=False),
                new_data_json=json.dumps({"status": "TERMINATED"}, ensure_ascii=False),
                detection_reason=f"{salary_month}月考勤表中未发现该员工",
                source_batch_id=batch_id,
                status="PENDING",
            )
            db.add(cand)
            possible_terminations += 1
            candidates.append({"type": "POSSIBLE_TERMINATION", "employee_name": emp.employee_name})

        db.commit()

        logger.info(
            "考勤比对完成: matched=%d, new=%d, terminations=%d, dept_changes=%d, conflicts=%d, summaries=%d, empty=%d",
            matched_count, new_hire_count, possible_terminations,
            sum(1 for c in candidates if c["type"] in ("DEPARTMENT_CHANGE", "POSITION_CHANGE")),
            sum(1 for c in candidates if c["type"] == "INFO_CONFLICT"),
            summary_count, empty_count,
        )

        return {
            "batch_id": batch_id,
            "matched_count": matched_count,
            "new_hire_count": new_hire_count,
            "possible_terminations": possible_terminations,
            "dept_changes": sum(1 for c in candidates if c["type"] == "DEPARTMENT_CHANGE"),
            "pos_changes": sum(1 for c in candidates if c["type"] == "POSITION_CHANGE"),
            "conflicts": sum(1 for c in candidates if c["type"] == "INFO_CONFLICT"),
            "summary_count": summary_count,
            "empty_count": empty_count,
            "total_candidates": len(candidates),
            "candidates": candidates,
        }

    except BizError:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.error("考勤导入失败: %s", str(e)[:300], exc_info=True)
        raise BizError(error_code="EXCEL_PARSE_FAILED", message="考勤导入失败，请确认文件格式正确。")
    finally:
        db.close()


def list_change_candidates(status: str | None = None, candidate_type: str | None = None, page: int = 1, page_size: int = 50) -> dict:
    """查询异动待确认记录"""
    db = SessionLocal()
    try:
        query = db.query(EmployeeChangeCandidate)
        if status and status != "ALL":
            query = query.filter(EmployeeChangeCandidate.status == status)
        if candidate_type and candidate_type != "ALL":
            query = query.filter(EmployeeChangeCandidate.candidate_type == candidate_type)
        query = query.order_by(EmployeeChangeCandidate.created_at.desc())
        total = query.count()
        items = query.offset((page - 1) * page_size).limit(page_size).all()
        return {
            "items": [
                {
                    "id": c.id,
                    "salary_month": c.salary_month,
                    "employee_master_id": c.employee_master_id,
                    "candidate_type": c.candidate_type,
                    "employee_no": c.employee_no or "",
                    "employee_name": c.employee_name,
                    "department": c.department or "",
                    "position": c.position or "",
                    "old_data": json.loads(c.old_data_json or "{}"),
                    "new_data": json.loads(c.new_data_json or "{}"),
                    "detection_reason": c.detection_reason or "",
                    "source_batch_id": c.source_batch_id or "",
                    "status": c.status,
                    "handled_by": c.handled_by or "",
                    "handle_remark": c.handle_remark or "",
                    "created_at": c.created_at.isoformat() if c.created_at else "",
                }
                for c in items
            ],
            "total": total,
        }
    finally:
        db.close()


def handle_change_candidate(candidate_id: str, action: str, handle_data: dict, handled_by: str) -> dict:
    """处理异动待确认记录"""
    db = SessionLocal()
    try:
        cand = db.query(EmployeeChangeCandidate).filter(
            EmployeeChangeCandidate.id == candidate_id
        ).first()
        if not cand:
            raise BizError(error_code="RESOURCE_NOT_FOUND", message="异动记录不存在")
        if cand.status != "PENDING":
            raise BizError(error_code="INVALID_ARGUMENT", message=f"异动记录已处理: {cand.status}")

        if action == "CONFIRM_HIRE":
            # 确认入职
            emp_id = f"emp_{uuid.uuid4().hex[:8]}"
            today = date.today()
            suggested_hire = cand.hire_date or date.today()
            emp = EmployeeMaster(
                id=emp_id,
                employee_no=_unique_employee_no(db, cand.employee_no or ""),
                employee_name=cand.employee_name,
                department=cand.department or handle_data.get("department", ""),
                position=cand.position or handle_data.get("position", ""),
                hire_date=_parse_date(handle_data.get("hire_date")) or suggested_hire,
                status="ACTIVE",
                created_by=handled_by,
            )
            db.add(emp)
            cand.status = "CONFIRMED"
            cand.employee_master_id = emp_id
            cand.handled_by = handled_by
            cand.handled_at = datetime.now()
            cand.handle_remark = handle_data.get("remark", "确认入职")

        elif action == "CONFIRM_TERMINATION":
            # 确认离职
            emp = db.query(EmployeeMaster).filter(EmployeeMaster.id == cand.employee_master_id).first()
            if not emp:
                raise BizError(error_code="RESOURCE_NOT_FOUND", message="员工档案不存在")
            term_date = _parse_date(handle_data.get("termination_date")) or date.today()
            emp.status = "RESIGNED"
            emp.resign_date = term_date
            # 记录任职历史
            _record_assignment(db, emp.id, emp.department, emp.position, term_date, "RESIGN",
                               handle_data.get("reason", "离职"), cand.source_batch_id, handled_by)
            cand.status = "CONFIRMED"
            cand.handled_by = handled_by
            cand.handled_at = datetime.now()
            cand.handle_remark = handle_data.get("remark", "确认离职")

        elif action == "CONFIRM_TRANSFER":
            # 确认调岗
            emp = db.query(EmployeeMaster).filter(EmployeeMaster.id == cand.employee_master_id).first()
            if not emp:
                raise BizError(error_code="RESOURCE_NOT_FOUND", message="员工档案不存在")
            old_dept = emp.department
            old_pos = emp.position
            if cand.department:
                emp.department = cand.department
            if cand.position:
                emp.position = cand.position
            _record_assignment(db, emp.id, emp.department, emp.position, date.today(), "TRANSFER",
                               handle_data.get("reason", f"部门: {old_dept}→{emp.department}, 岗位: {old_pos}→{emp.position}"),
                               cand.source_batch_id, handled_by)
            cand.status = "CONFIRMED"
            cand.handled_by = handled_by
            cand.handled_at = datetime.now()
            cand.handle_remark = handle_data.get("remark", "确认调岗")

        elif action == "IGNORE":
            cand.status = "IGNORED"
            cand.handled_by = handled_by
            cand.handled_at = datetime.now()
            cand.handle_remark = handle_data.get("remark", "忽略")

        elif action == "KEEP_ACTIVE":
            cand.status = "REJECTED"
            cand.handled_by = handled_by
            cand.handled_at = datetime.now()
            cand.handle_remark = "员工在职，忽略离职提醒"

        db.commit()
        return {"id": cand.id, "status": cand.status}
    except BizError:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.error("处理异动失败: %s", str(e)[:200], exc_info=True)
        raise BizError(error_code="INTERNAL_ERROR", message="处理异动记录失败")
    finally:
        db.close()


def batch_confirm(action: str, candidate_ids: list[str], handle_data: dict, handled_by: str) -> dict:
    """批量确认异动"""
    confirmed = 0
    errors = []
    for cid in candidate_ids:
        try:
            handle_change_candidate(cid, action, handle_data, handled_by)
            confirmed += 1
        except Exception as e:
            errors.append({"id": cid, "error": str(e)})
    return {"confirmed": confirmed, "errors": errors}


def get_candidate_counts() -> dict:
    """获取异动待确认统计"""
    db = SessionLocal()
    try:
        rows = db.query(
            EmployeeChangeCandidate.candidate_type,
            EmployeeChangeCandidate.status,
        ).all()
        counts = {"POSSIBLE_HIRE": 0, "POSSIBLE_TERMINATION": 0, "DEPARTMENT_CHANGE": 0,
                  "POSITION_CHANGE": 0, "INFO_CONFLICT": 0, "PENDING_TOTAL": 0}
        for cand_type, status in rows:
            if status == "PENDING" and cand_type in counts:
                counts[cand_type] += 1
                counts["PENDING_TOTAL"] += 1
        return counts
    finally:
        db.close()


# ============================================================
#  内部工具
# ============================================================

def _detect_col(headers: list, keywords: list[str]) -> int | None:
    for i, h in enumerate(headers):
        if h:
            hs = str(h).strip()
            for kw in keywords:
                if kw in hs:
                    return i
    return None


def _find_employee(db, employee_no: str, employee_name: str) -> EmployeeMaster | None:
    if employee_no:
        emp = db.query(EmployeeMaster).filter(
            EmployeeMaster.employee_no == employee_no,
        ).first()
        if emp:
            return emp
    from app.services.pool_service import clean_name
    clean = clean_name(employee_name)
    if clean:
        emps = db.query(EmployeeMaster).filter(
            EmployeeMaster.employee_name == clean,
        ).all()
        if len(emps) == 1:
            return emps[0]
    return None


def _record_assignment(db, emp_id, department, position, effective_date, change_type, reason, batch_id, created_by):
    ah = EmployeeAssignmentHistory(
        id=f"ah_{uuid.uuid4().hex[:8]}",
        employee_master_id=emp_id,
        department=department or "",
        position=position or "",
        effective_start_date=effective_date,
        change_type=change_type,
        change_reason=reason or "",
        source_batch_id=batch_id or "",
        created_by=created_by,
    )
    db.add(ah)


def _parse_date(val) -> date | None:
    if not val:
        return None
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val)[:10], "%Y-%m-%d").date()
    except Exception:
        return None
