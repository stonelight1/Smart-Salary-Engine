"""
外部引用版本服务

负责：
- 解析上传的上月最终工资表 Excel
- 自动识别字段（姓名、编号、部门、岗位、工资标准等）
- 过滤汇总行、空白行
- 保存引用版本快照
- 生成本月工资草稿
"""

import hashlib
import json
import logging
import os
import tempfile
import uuid
from decimal import Decimal
from pathlib import Path

from app.core.exceptions import BizError
from app.db.database import SessionLocal
from app.models import SalaryRun, SalaryReferenceSource, EmployeeRecord
from app.services.attendance_service import get_attendance_rule
from app.engines.importer.importer import validate_file, save_file, parse_workbook
from app.services.pool_service import classify_row_type, clean_name, SUMMARY_KEYWORDS
from app.db.database import PROJECT_ROOT

logger = logging.getLogger(__name__)

UPLOADS_DIR = PROJECT_ROOT / "uploads" / "reference_excel"
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

# 期望识别的字段别名映射
FIELD_ALIASES: dict[str, list[str]] = {
    "employee_name": ["姓名", "员工姓名", "人员姓名", "名字", "员工", "姓名"],
    "employee_no": ["员工编号", "工号", "编号", "员工ID", "员工工号"],
    "department": ["部门", "所属部门", "组织", "科室", "部门名称"],
    "position": ["岗位", "职位", "职务", "岗位名称"],
    "salary_standard": ["工资标准", "月薪标准", "满绩效月薪", "薪资标准", "全额工资", "月薪总额", "合同工资", "月薪", "标准工资", "月薪合计"],
    "basic_salary": ["基本工资", "基本薪资", "基础工资", "底薪"],
    "performance_salary_standard": ["标准绩效", "绩效基数", "绩效工资标准", "标准绩效工资", "绩效标准", "绩效总额", "绩效工资"],
    "social_security_personal": ["社保个人", "社保扣款", "养老保险", "医疗保险", "失业保险"],
    "housing_fund_personal": ["公积金个人", "公积金扣款", "住房公积金"],
}
# 继承字段（从引用导入本月草稿）
INHERIT_FIELDS = {
    "employee_name", "employee_no", "department", "position",
    "salary_standard", "basic_salary", "performance_salary_standard",
    "social_security_personal", "housing_fund_personal",
}


def _parse_with_calamine(file_path: str) -> dict:
    """
    使用 python-calamine 兼容模式解析 Excel（适用于 openpyxl 无法读取的损坏文件）。

    返回与 parse_workbook 兼容的结构，额外包含 _raw_data 和 _raw_rows。
    """
    try:
        from python_calamine import CalamineWorkbook
    except ImportError:
        raise BizError(
            error_code="EXCEL_REPAIR_FAILED",
            message="系统未安装兼容 Excel 解析组件，请使用 Excel 打开文件并另存为 .xlsx 后重新上传。",
        )

    try:
        wb = CalamineWorkbook.from_path(file_path)
    except Exception as e:
        raise BizError(
            error_code="EXCEL_PARSE_FAILED",
            message=f"Excel 文件无法识别，请确认文件格式正确。",
        )

    sheet_names = wb.sheet_names
    if not sheet_names:
        raise BizError(error_code="EXCEL_PARSE_FAILED", message="Excel 文件中未找到任何工作表")

    sheets = []
    total_rows = 0

    for sname in sheet_names:
        try:
            sheet = wb.get_sheet_by_name(sname)
            rows = sheet.to_python()
        except Exception as e:
            continue

        if not rows:
            continue

        total_rows += len(rows)

        # 查找表头行（扫描前 10 行）
        header_idx = None
        header_values = []
        for i, row in enumerate(rows[:10]):
            non_empty = [v for v in row if v is not None and str(v).strip()]
            if len(non_empty) < 2:
                continue
            vals = [str(v).strip() for v in non_empty]
            known_kw = ["姓名", "员工", "工资", "基本", "部门", "金额", "出勤", "绩效", "扣款", "社保", "补贴"]
            hits = sum(1 for v in vals if any(kw in v for kw in known_kw))
            if hits >= 1:
                header_idx = i
                header_values = [str(v).strip() for v in non_empty]
                break

        if header_idx is None:
            header_values = [str(v).strip() for v in rows[0] if v is not None]
            header_idx = 0

        # 合并跨行列标题 - 当第2行存在薪资关键词时合并
        merged = False
        for i in range(header_idx + 1, min(header_idx + 3, len(rows))):
            next_row = rows[i]
            next_vals = [str(v).strip() for v in next_row[:15] if v is not None and str(v).strip()]
            kw = ["基本", "绩效", "补贴", "扣款", "社保", "加班", "应发", "实发", "小计", "合计"]
            hits = sum(1 for v in next_vals if any(k in v for k in kw))
            if hits >= 2:
                top = list(header_values)
                for j in range(min(len(next_row), 15)):
                    sub_val = str(next_row[j]).strip() if next_row[j] is not None else ""
                    if sub_val:
                        if j < len(top):
                            # 子标题有薪资关键词时，替换父标题
                            salary_kw = ["基本", "绩效", "社保", "补贴", "加班", "扣款"]
                            is_salary_sub = any(k in sub_val for k in salary_kw)
                            top_has_salary = any(k in top[j] for k in salary_kw) if top[j] else False
                            if is_salary_sub and not top_has_salary:
                                top[j] = sub_val
                        else:
                            top.append(sub_val)
                header_values = top
                merged = True
                break

        data_start_idx = header_idx + (2 if merged else 1)
        preview = rows[data_start_idx:data_start_idx + 5] if data_start_idx < len(rows) else []

        sheets.append({
            "sheet_name": sname,
            "row_count": len(rows),
            "header_row_index": data_start_idx - 1 if data_start_idx > 0 else 0,
            "headers": header_values,
            "preview_rows": preview,
        })

    if not sheets:
        raise BizError(error_code="EXCEL_PARSE_FAILED", message="文件中没有有效的工作表数据")

    # 将所有行数据存储为字典值
    all_rows_data = {}
    for sname in sheet_names:
        try:
            sheet = wb.get_sheet_by_name(sname)
            all_rows_data[sname] = sheet.to_python()
        except Exception:
            all_rows_data[sname] = []

    return {
        "sheet_count": len(sheets),
        "sheets": sheets,
        "total_rows": total_rows,
        "_raw_data": all_rows_data.get(sheets[0]["sheet_name"], []),
        "_raw_rows": all_rows_data,
    }


def parse_reference_excel(file_bytes: bytes, original_name: str) -> dict:
    """
    解析上传的上月最终工资表 Excel。

    Returns:
        {
            "success": True,
            "file_path": "...",
            "file_hash": "...",
            "sheet_name": "...",
            "field_mapping": {...},
            "employees": [{...}],
            "employee_count": N,
            "summary_row_count": N,
            "empty_row_count": N,
            "invalid_row_count": N,
            "issues": [{...}],
        }
    """
    validate_file(file_bytes, original_name)
    storage_path, file_hash = save_file(file_bytes, original_name, "ref")

    # 尝试标准解析（openpyxl），失败时使用 calamine 降级
    parse_result = None
    used_calamine = False
    try:
        parse_result = parse_workbook(storage_path)
    except BizError as e:
        if "自动修复" in e.message:
            # openpyxl 失败 → 尝试 calamine
            parse_result = _parse_with_calamine(storage_path)
            used_calamine = True
        else:
            raise

    if parse_result is None:
        parse_result = _parse_with_calamine(storage_path)
        used_calamine = True

    if parse_result["sheet_count"] == 0:
        raise BizError(error_code="EXCEL_PARSE_FAILED", message="Excel 文件中未找到任何工作表")

    sheet = parse_result["sheets"][0]
    sheet_name = sheet["sheet_name"]
    headers = sheet["headers"]
    header_row = sheet["header_row_index"]

    if not headers:
        raise BizError(error_code="EXCEL_PARSE_FAILED", message="未识别到表头行，请确保 Excel 第一行为列标题")

    # 自动识别字段映射
    field_mapping = _auto_detect_mapping(headers)
    if "employee_name" not in field_mapping:
        raise BizError(
            error_code="EXCEL_PARSE_FAILED",
            message="未识别到员工姓名字段，请确保 Excel 包含「姓名」或「员工姓名」列",
        )

    # 读取原始数据
    all_cells = None
    if used_calamine:
        all_values = parse_result["_raw_data"]
        all_rows_data = parse_result["_raw_rows"]
    else:
        from openpyxl import load_workbook
        wb_values = load_workbook(storage_path, data_only=True)
        ws_values = wb_values[sheet_name]
        all_values = list(ws_values.iter_rows(values_only=True))
        wb_values.close()

        # 也加载 cell 对象用于公式检测
        wb_cells = load_workbook(storage_path, data_only=False)
        ws_cells = wb_cells[sheet_name]
        all_cells = list(ws_cells.iter_rows(values_only=False))
        wb_cells.close()

    data_start = (header_row or 0) + 1

    # 构建行内字段→列索引映射
    name_idx = field_mapping.get("employee_name")
    field_to_col = {}
    for field, header_val in field_mapping.items():
        if header_val in headers:
            field_to_col[field] = headers.index(header_val)

    # 姓名列索引（用于 calamine 降级）
    emp_name_col = field_to_col.get("employee_name")

    employees = []
    issues = []
    summary_count = 0
    empty_count = 0
    invalid_count = 0
    seen_names = {}
    seen_emp_nos = {}

    # 行关键词（用于 calamine 降级时的检测）
    summary_keywords = ["合计", "总计", "汇总", "小计", "部门合计", "总结", "自动总结", "总汇总"]

    for idx in range(data_start, len(all_values)):
        val_row = all_values[idx]

        # 行分类
        if used_calamine:
            # calamine 降级：基于原始值判断
            row_values = [v for v in val_row if v is not None and str(v).strip()]
            name_cell = val_row[emp_name_col] if emp_name_col < len(val_row) else None
            name_val_raw = str(name_cell).strip() if name_cell is not None else ""

            # 检查汇总关键词
            is_summary = False
            for v in val_row:
                if v is not None and isinstance(v, str):
                    for kw in summary_keywords:
                        if kw in v:
                            is_summary = True
                            break
                if is_summary:
                    break

            if is_summary:
                summary_count += 1
                continue
            if not row_values:
                empty_count += 1
                continue
            if not name_val_raw:
                # 姓名为空但有数据
                has_other_data = False
                for ci, v in enumerate(val_row):
                    if v is not None and ci != emp_name_col:
                        has_other_data = True
                        break
                if has_other_data:
                    invalid_count += 1
                else:
                    empty_count += 1
                continue
            row_type = "EMPLOYEE"
        else:
            cell_row = list(all_cells[idx])
            row_type = classify_row_type(cell_row, emp_name_col, field_to_col)
            if row_type == "SUMMARY":
                summary_count += 1
                continue
            if row_type == "EMPTY":
                empty_count += 1
                continue
            if row_type == "INVALID":
                invalid_count += 1
                continue

        # EMPLOYEE
        emp = _extract_employee(val_row, field_to_col, field_mapping, idx + 1)
        if not emp["employee_name"]:
            invalid_count += 1
            continue

        # 重名检测
        name_clean = clean_name(emp["employee_name"])
        if name_clean in seen_names:
            prev = seen_names[name_clean]
            issues.append({
                "type": "DUPLICATE_NAME",
                "row": idx + 1,
                "message": f"员工姓名重复：{name_clean}（第{prev}行和第{idx+1}行）",
                "level": "BLOCK",
            })
        seen_names[name_clean] = idx + 1

        # 员工编号唯一性
        emp_no = emp.get("employee_no", "")
        if emp_no and emp_no in seen_emp_nos:
            prev_row = seen_emp_nos[emp_no]
            prev_name = next((e["employee_name"] for e in employees if e.get("employee_no") == emp_no), "")
            issues.append({
                "type": "CONFLICT_EMP_NO",
                "row": idx + 1,
                "message": f"员工编号冲突：{emp_no}（第{prev_row}行「{prev_name}」和第{idx+1}行「{name_clean}」）",
                "level": "BLOCK",
            })
        if emp_no:
            seen_emp_nos[emp_no] = idx + 1

        employees.append(emp)

    return {
        "success": True,
        "file_path": storage_path,
        "file_hash": file_hash,
        "sheet_name": sheet_name,
        "field_mapping": field_mapping,
        "employees": employees,
        "employee_count": len(employees),
        "summary_row_count": summary_count,
        "empty_row_count": empty_count,
        "invalid_row_count": invalid_count,
        "issues": issues,
    }


def create_reference_source(
    file_bytes: bytes,
    original_name: str,
    parsed: dict,
    target_salary_month: str,
    reference_salary_month: str,
    created_by: str,
) -> dict:
    """
    保存外部引用版本到数据库。

    Returns:
        {"reference_id": "...", "status": "...", "employee_count": N}
    """
    db = SessionLocal()
    try:
        ref_id = f"ref_{uuid.uuid4().hex[:8]}"
        file_hash = hashlib.sha256(file_bytes).hexdigest()

        # 保存到引用目录
        ref_dir = UPLOADS_DIR / ref_id
        ref_dir.mkdir(parents=True, exist_ok=True)
        saved_path = ref_dir / original_name
        with open(saved_path, "wb") as f:
            f.write(file_bytes)

        # 序列化员工数据（不含敏感字段）
        employees_summary = []
        for emp in parsed["employees"]:
            employees_summary.append({
                "employee_no": emp.get("employee_no", ""),
                "employee_name": emp.get("employee_name", ""),
                "department": emp.get("department", ""),
                "position": emp.get("position", ""),
                "salary_standard": str(emp.get("salary_standard", "")),
                "basic_salary": str(emp.get("basic_salary", "")),
                "performance_salary_standard": str(emp.get("performance_salary_standard", "")),
            })

        ref = SalaryReferenceSource(
            id=ref_id,
            source_type="EXTERNAL_EXCEL",
            reference_salary_month=reference_salary_month,
            target_salary_month=target_salary_month,
            original_file_name=original_name,
            file_storage_path=str(saved_path),
            file_hash=file_hash,
            file_size=len(file_bytes),
            sheet_name=parsed["sheet_name"],
            status="PENDING_CONFIRM",
            employee_count=parsed["employee_count"],
            summary_row_count=parsed["summary_row_count"],
            empty_row_count=parsed["empty_row_count"],
            invalid_row_count=parsed["invalid_row_count"],
            field_mapping_json=json.dumps(parsed["field_mapping"], ensure_ascii=False),
            parsed_data_json=json.dumps(employees_summary, ensure_ascii=False),
            created_by=created_by,
        )
        db.add(ref)
        db.commit()

        return {
            "reference_id": ref_id,
            "status": "PENDING_CONFIRM",
            "employee_count": parsed["employee_count"],
        }
    except BizError:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.error("保存引用版本失败: %s", str(e)[:300], exc_info=True)
        raise BizError(error_code="INTERNAL_ERROR", message="保存引用版本失败")
    finally:
        db.close()


def confirm_reference_and_create_draft(
    reference_id: str,
    target_run_id: str,
    created_by: str,
) -> dict:
    """
    确认引用版本，并使用引用的员工数据创建本月工资草稿。
    """
    db = SessionLocal()
    try:
        ref = db.query(SalaryReferenceSource).filter(SalaryReferenceSource.id == reference_id).first()
        if not ref:
            raise BizError(error_code="RESOURCE_NOT_FOUND", message="引用版本不存在")
        if ref.status != "PENDING_CONFIRM":
            raise BizError(error_code="INVALID_ARGUMENT", message=f"引用版本状态不正确: {ref.status}")

        target_run = db.query(SalaryRun).filter(SalaryRun.id == target_run_id).first()
        if not target_run:
            raise BizError(error_code="RESOURCE_NOT_FOUND", message="目标任务不存在")

        target_month = target_run.payroll_month
        employees_data = json.loads(ref.parsed_data_json or "[]")

        copied_count = 0
        for emp_data in employees_data:
            emp_id = f"emp_{uuid.uuid4().hex[:8]}"
            salary_std = _parse_decimal(emp_data.get("salary_standard"))
            basic_sal = _parse_decimal(emp_data.get("basic_salary"))
            perf_std = _parse_decimal(emp_data.get("performance_salary_standard"))

            dept = emp_data.get("department", "") or ""
            pos = emp_data.get("position", "") or ""
            rule = get_attendance_rule(dept, pos, target_month)

            emp = EmployeeRecord(
                id=emp_id,
                salary_run_id=target_run_id,
                employee_name=emp_data.get("employee_name", ""),
                status="NORMAL",
                employee_no=emp_data.get("employee_no", "") or None,
                department=dept or None,
                position=pos or None,
                salary_standard=salary_std,
                basic_salary=basic_sal,
                performance_salary_standard=perf_std,
                performance_score=None,
                performance_score_source=None,
                performance_salary_actual=None,
                attendance_rule_type=rule["rule_type"],
                attendance_rule_name=rule["rule_name"],
                standard_attendance_days=rule["standard_attendance_days"],
                leave_days=None,
                actual_attendance_days=None,
                attendance_deduction=None,
                gross_salary_before_attendance=None,
            )
            db.add(emp)
            copied_count += 1

        ref.status = "CONFIRMED"
        ref.confirmed_by = created_by
        ref.confirmed_at = __import__("datetime").datetime.now()

        target_run.reference_external_id = reference_id
        target_run.reference_source_type = "EXTERNAL_EXCEL"
        target_run.status = "IMPORTED"
        db.commit()

        logger.info("外部引用确认+草稿生成: ref=%s, target=%s, employees=%d", reference_id, target_run_id, copied_count)

        return {
            "status": "DRAFT",
            "employee_count": copied_count,
        }
    except BizError:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.error("确认引用失败: %s", str(e)[:300], exc_info=True)
        raise BizError(error_code="INTERNAL_ERROR", message="确认引用并生成本月草稿失败")
    finally:
        db.close()


def _auto_detect_mapping(headers: list[str]) -> dict[str, str]:
    """
    自动识别表头到系统字段的映射。
    Returns: {field_code: original_header_string}
    """
    mapping: dict[str, str] = {}
    used_headers = set()

    for field_code, aliases in FIELD_ALIASES.items():
        best_match = None
        best_score = 0
        for header in headers:
            if header in used_headers:
                continue
            h_str = str(header).strip()
            for alias in aliases:
                if alias == h_str:
                    best_match = h_str
                    best_score = 1.0
                    break
                if alias in h_str:
                    score = len(alias) / max(len(h_str), 1)
                    if score > best_score:
                        best_score = score
                        best_match = h_str

        if best_match and best_score >= 0.5:
            mapping[field_code] = best_match
            used_headers.add(best_match)

    return mapping


def _extract_employee(val_row: tuple, field_to_col: dict, field_mapping: dict, row_num: int) -> dict:
    """从一行数据中提取员工字段"""
    emp: dict[str, object] = {"_row": row_num}

    for field_code in ["employee_name", "employee_no", "department", "position",
                        "salary_standard", "basic_salary", "performance_salary_standard"]:
        col_idx = field_to_col.get(field_code)
        if col_idx is not None and col_idx < len(val_row):
            raw = val_row[col_idx]
            if raw is not None and str(raw).strip():
                if field_code in ("salary_standard", "basic_salary", "performance_salary_standard"):
                    emp[field_code] = _parse_decimal(raw)
                else:
                    emp[field_code] = str(raw).strip()
            else:
                emp[field_code] = "" if field_code in ("employee_name", "employee_no", "department", "position") else None
        else:
            emp[field_code] = "" if field_code in ("employee_name", "employee_no", "department", "position") else None

    name_clean = clean_name(str(emp.get("employee_name", "")))
    emp["employee_name"] = name_clean

    return emp


def _parse_decimal(value) -> Decimal | None:
    if value is None:
        return None
    try:
        text = str(value).replace(",", "").replace("￥", "").replace("¥", "").replace(" ", "").strip()
        if text in ("", "-", "—", "--", "null", "None"):
            return None
        return Decimal(text)
    except Exception:
        return None
