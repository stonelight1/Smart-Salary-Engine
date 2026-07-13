"""员工数据池服务

负责：
- 按姓名清洗和聚合员工数据
- 保存字段值和来源
- 多批次补充数据合并
- 检测数据冲突和姓名重复
- 行类型分类（过滤汇总行、空白行）
"""

import uuid
from decimal import Decimal
from typing import Any

from app.core.exceptions import BizError
from app.db.database import SessionLocal
from app.models import EmployeeRecord, EmployeeFieldValue, CheckIssue, ImportBatch, SheetMapping, ColumnMapping
from app.engines.importer.importer import parse_workbook
from openpyxl import load_workbook

SUMMARY_KEYWORDS = ["合计", "总计", "汇总", "小计"]

# ============================================================
#  行类型分类
# ============================================================

def classify_row_type(
    cell_row: list,
    name_idx: int | None,
    field_to_col: dict[str, int],
) -> str:
    """
    对一行数据进行分类。

    Returns:
        'EMPLOYEE' — 有效员工数据行
        'SUMMARY'  — 合计/汇总行，跳过
        'EMPTY'    — 空白行，跳过
        'INVALID'  — 有姓名但缺少关键字段，不创建员工
    """
    # 1. 完全空行
    if not cell_row or all(c.value is None for c in cell_row):
        return "EMPTY"

    # 2. 关键词检测：任意单元格包含合计/汇总等关键词
    for c in cell_row:
        if c.value is not None and isinstance(c.value, str):
            val = str(c.value).strip()
            for kw in SUMMARY_KEYWORDS:
                if kw in val:
                    return "SUMMARY"

    # 3. 提取姓名
    name_clean = ""
    if name_idx is not None and name_idx < len(cell_row):
        name_cell = cell_row[name_idx]
        if name_cell.value is not None:
            name_clean = clean_name(str(name_cell.value))

    # 4. 公式检测：无姓名时检查是否有 SUM/SUBTOTAL 公式
    if not name_clean:
        has_formula_sum = False
        has_non_empty_field = False
        for field_code, col_idx in field_to_col.items():
            if col_idx >= len(cell_row):
                continue
            cell = cell_row[col_idx]
            # 检测 SUM/SUBTOTAL 公式
            if cell.data_type == "f":
                formula = str(cell.value or "").upper()
                if "SUM" in formula or "SUBTOTAL" in formula:
                    has_formula_sum = True
            # 检测任何字段的非空值（非姓名列）
            if cell.value is not None and col_idx != name_idx:
                has_non_empty_field = True

        if has_formula_sum:
            return "SUMMARY"

        if not has_non_empty_field:
            return "EMPTY"

        # 姓名列为空但有其他非空字段 → 视为无效/跳过
        return "EMPTY"

    # 5. 有姓名，验证关键辅助字段
    supporting_fields = ["employee_no", "department", "position", "base_salary"]
    has_support = any(
        col_idx is not None and col_idx < len(cell_row) and cell_row[col_idx].value is not None
        for field_code in supporting_fields
        if (col_idx := field_to_col.get(field_code)) is not None
    )

    if not has_support:
        return "INVALID"

    return "EMPLOYEE"


def clean_name(name: str) -> str:
    """清洗员工姓名"""
    if not name:
        return ""
    name = name.strip()
    name = name.replace("　", " ")   # 全角空格
    name = name.replace(" ", " ")   # 非断行空格
    return name.strip()


# ============================================================
#  数据池构建
# ============================================================

def build_data_pool(salary_run_id: str, import_batch_id: str) -> dict[str, Any]:
    """从最新导入批次构建/更新员工数据池"""
    db = SessionLocal()
    try:
        # 1. 获取导入批次的文件信息
        batch = db.query(ImportBatch).filter(ImportBatch.id == import_batch_id).first()
        if not batch:
            raise BizError(error_code="RESOURCE_NOT_FOUND", message="导入批次不存在")

        from app.models import WorkbookFile
        wf = db.query(WorkbookFile).filter(WorkbookFile.id == batch.file_id).first()
        if not wf:
            raise BizError(error_code="RESOURCE_NOT_FOUND", message="文件记录不存在")

        # 2. 获取列映射（标准字段 → 原始列索引）
        sheet_mappings = db.query(SheetMapping).filter(
            SheetMapping.import_batch_id == import_batch_id
        ).all()

        all_employees = []
        new_count = 0
        duplicate_count = 0
        total_summary_count = 0
        total_empty_count = 0
        total_invalid_count = 0
        total_employee_count = 0

        for sheet in parse_workbook(wf.storage_path)["sheets"]:
            sheet_name = sheet["sheet_name"]
            headers = sheet["headers"]
            header_row = sheet["header_row_index"]

            # 找对应的 sheet_mapping
            sm = next((s for s in sheet_mappings if s.sheet_name == sheet_name), None)
            if not sm or sm.sheet_type in ("unknown",):
                continue

            # 构建列索引映射
            col_mappings = db.query(ColumnMapping).filter(
                ColumnMapping.sheet_mapping_id == sm.id
            ).all()

            field_to_col: dict[str, int] = {}
            for cm in col_mappings:
                if cm.field_code and cm.original_column in headers:
                    field_to_col[cm.field_code] = headers.index(cm.original_column)

            if "employee_name" not in field_to_col:
                continue  # 没有姓名列，跳过

            # 获取字段配置中的类型信息
            from app.core.config import config_loader
            fields_config = config_loader.get("fields") or {}
            field_defs = fields_config.get("fields", {})

            # 直接从 Excel 读取
            # 分别加载：formula 模式用于公式检测，value 模式用于数据值
            wb_formula = load_workbook(wf.storage_path, data_only=False)
            ws_formula = wb_formula[sheet_name]
            all_cells = list(ws_formula.iter_rows(values_only=False))
            wb_formula.close()

            wb_values = load_workbook(wf.storage_path, data_only=True)
            ws_values = wb_values[sheet_name]
            all_values = list(ws_values.iter_rows(values_only=True))
            wb_values.close()

            data_start = (header_row or 0) + 1
            name_col = field_to_col["employee_name"]

            employee_rows: list[dict[str, Any]] = []

            for idx in range(data_start, len(all_cells)):
                cell_row = list(all_cells[idx])
                val_row = all_values[idx] if idx < len(all_values) else tuple()

                row_type = classify_row_type(cell_row, name_col, field_to_col)

                if row_type == "SUMMARY":
                    total_summary_count += 1
                    continue
                if row_type == "EMPTY":
                    total_empty_count += 1
                    continue
                if row_type == "INVALID":
                    total_invalid_count += 1
                    continue

                # EMPLOYEE 行
                total_employee_count += 1

                # 提取姓名
                name_raw = str(val_row[name_col]) if name_col < len(val_row) and val_row[name_col] is not None else ""
                name_clean = clean_name(name_raw)

                if not name_clean:
                    # 理论不应到达此处（classify 已检查），防御性处理
                    total_invalid_count += 1
                    continue

                # 检查重名
                existing = db.query(EmployeeRecord).filter(
                    EmployeeRecord.salary_run_id == salary_run_id,
                    EmployeeRecord.employee_name == name_clean,
                ).first()

                if existing:
                    # 检查是否是同一批次内重名
                    if existing.status == "NAME_DUPLICATE":
                        duplicate_count += 1
                        continue

                    # 检查是否在同一批次文件内重名
                    dup_check = db.query(EmployeeFieldValue).filter(
                        EmployeeFieldValue.employee_record_id == existing.id,
                        EmployeeFieldValue.import_batch_id == import_batch_id,
                    ).first()

                    if dup_check:
                        # 同一批次内重名 → BLOCK
                        existing.status = "NAME_DUPLICATE"
                        issue = CheckIssue(
                            id=f"issue_{uuid.uuid4().hex[:8]}",
                            salary_run_id=salary_run_id,
                            employee_record_id=existing.id,
                            issue_code="EMPLOYEE_DUPLICATED",
                            level="BLOCK",
                            message=f"发现同名员工: {name_clean}，请检查 Excel 后重新导入",
                            status="OPEN",
                        )
                        db.add(issue)
                        duplicate_count += 1
                        db.commit()
                        continue

                    # 不同导入批次间的同名 → 合并数据
                    _merge_employee_fields(
                        db, existing.id, import_batch_id,
                        field_to_col, val_row, field_defs,
                        wf.id, sheet_name, batch.file_role,
                    )
                    continue

                # 新员工
                emp_id = f"emp_{uuid.uuid4().hex[:8]}"
                emp = EmployeeRecord(
                    id=emp_id,
                    salary_run_id=salary_run_id,
                    employee_name=name_clean,
                    status="NORMAL",
                )
                db.add(emp)

                # 保存字段值
                for field_code, col_idx in field_to_col.items():
                    if col_idx >= len(val_row):
                        continue
                    raw_value = val_row[col_idx]
                    if raw_value is None:
                        continue

                    field_info = field_defs.get(field_code, {})
                    value_type = field_info.get("type", "string")
                    value_text = str(raw_value)
                    value_decimal = _to_decimal(raw_value) if value_type == "money" or value_type == "number" else None

                    fv = EmployeeFieldValue(
                        id=f"fv_{uuid.uuid4().hex[:8]}",
                        employee_record_id=emp_id,
                        field_code=field_code,
                        value_text=value_text,
                        value_decimal=value_decimal,
                        value_type=value_type,
                        source_file_id=wf.id,
                        source_sheet=sheet_name,
                        source_row=idx + 1,
                        source_column=head_str(col_idx, headers),
                        import_batch_id=import_batch_id,
                        is_manual=False,
                    )
                    db.add(fv)

                new_count += 1
                all_employees.append(name_clean)
                db.commit()

        # 更新任务状态
        from app.models import SalaryRun
        run = db.query(SalaryRun).filter(SalaryRun.id == salary_run_id).first()
        if run and run.status in ("CREATED", "CHECK_FAILED"):
            run.status = "IMPORTED"
            db.commit()

        return {
            "new_count": new_count,
            "duplicate_count": duplicate_count,
            "total": db.query(EmployeeRecord).filter(
                EmployeeRecord.salary_run_id == salary_run_id,
            ).count(),
            "employee_row_count": total_employee_count,
            "summary_row_count": total_summary_count,
            "empty_row_count": total_empty_count,
            "invalid_row_count": total_invalid_count,
        }
    except BizError:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise BizError(
            error_code="INTERNAL_ERROR",
            message=f"数据池构建失败: {str(e)}",
        )
    finally:
        db.close()


def _merge_employee_fields(
    db: SessionLocal,
    emp_id: str,
    import_batch_id: str,
    field_to_col: dict[str, int],
    row_data: tuple,
    field_defs: dict,
    file_id: str,
    sheet_name: str,
    file_role: str,
):
    """合并已有员工的字段数据"""
    for field_code, col_idx in field_to_col.items():
        if col_idx >= len(row_data):
            continue
        raw_value = row_data[col_idx]
        if raw_value is None:
            continue

        # 检查该字段是否已有值
        existing_fv = db.query(EmployeeFieldValue).filter(
            EmployeeFieldValue.employee_record_id == emp_id,
            EmployeeFieldValue.field_code == field_code,
        ).first()

        if not existing_fv:
            # 新字段，直接添加
            field_info = field_defs.get(field_code, {})
            value_type = field_info.get("type", "string")
            value_text = str(raw_value)
            value_decimal = _to_decimal(raw_value) if value_type in ("money", "number") else None

            fv = EmployeeFieldValue(
                id=f"fv_{uuid.uuid4().hex[:8]}",
                employee_record_id=emp_id,
                field_code=field_code,
                value_text=value_text,
                value_decimal=value_decimal,
                value_type=value_type,
                source_file_id=file_id,
                source_sheet=sheet_name,
                source_row=0,
                source_column=field_code,
                import_batch_id=import_batch_id,
                is_manual=False,
            )
            db.add(fv)
        else:
            # 已有值，检查冲突
            new_text = str(raw_value).strip()
            old_text = (existing_fv.value_text or "").strip()
            if new_text and new_text != old_text:
                issue = CheckIssue(
                    id=f"issue_{uuid.uuid4().hex[:8]}",
                    salary_run_id=existing_fv.employee_record_id,
                    employee_record_id=emp_id,
                    issue_code="DATA_CONFLICT",
                    level="BLOCK",
                    field_code=field_code,
                    message=f"字段 {field_code} 数据冲突: 原值={old_text}, 新值={new_text}",
                    status="OPEN",
                )
                db.add(issue)


def _to_decimal(value: Any) -> Decimal | None:
    """尝试转换成 Decimal"""
    if value is None:
        return None
    try:
        text = str(value).replace(",", "").replace("￥", "").replace("¥", "").strip()
        if text in ("", "-", "—", "--"):
            return None
        return Decimal(text)
    except Exception:
        return None


def head_str(col_idx: int, headers: list[str]) -> str:
    """获取列名（由索引或表头决定）"""
    if col_idx < len(headers):
        return str(headers[col_idx])
    from openpyxl.utils import get_column_letter
    return get_column_letter(col_idx + 1)
