"""
绩效得分导入服务

负责：
- 解析绩效 Excel（员工姓名 + 绩效得分）
- 匹配员工记录
- 处理百分比/小数/文本格式
- 去重和异常检测
- 更新员工月度绩效得分
"""

import logging
import uuid
from decimal import Decimal, InvalidOperation

from app.core.exceptions import BizError
from app.db.database import SessionLocal
from app.models import (
    EmployeeRecord, PerformanceImportRecord, CheckIssue, SalaryRun,
)
from app.engines.importer.importer import parse_workbook, validate_file, save_file

logger = logging.getLogger(__name__)


def import_performance_scores(
    file_bytes: bytes,
    original_name: str,
    salary_run_id: str,
    created_by: str,
) -> dict:
    """
    导入绩效得分 Excel。

    Excel 格式：员工姓名（或员工编号）+ 绩效得分
    得分可以是 90%, 0.9, 90 等格式

    Returns:
        {
            "matched_count": int,
            "unmatched_names": [str],
            "duplicate_records": int,
            "error_records": [{...}],
            "batch_id": str,
        }
    """
    # 1. 校验
    validate_file(file_bytes, original_name)

    # 2. 保存文件
    storage_path, file_hash = save_file(file_bytes, original_name, salary_run_id)

    # 3. 解析 Excel
    parse_result = parse_workbook(storage_path)
    if parse_result["sheet_count"] == 0:
        raise BizError(error_code="EXCEL_PARSE_FAILED", message="未找到工作表")

    # 取第一个非空 Sheet
    sheet = parse_result["sheets"][0]
    headers = sheet["headers"]
    preview_rows = sheet["preview_rows"]
    header_row = sheet["header_row_index"] or 0

    # 检测哪列是姓名/编号，哪列是得分
    name_col = _detect_name_column(headers)
    score_col = _detect_score_column(headers)

    if name_col is None:
        raise BizError(
            error_code="EXCEL_PARSE_FAILED",
            message="未识别到员工姓名/编号列，请确保包含「姓名」或「员工编号」列",
        )
    if score_col is None:
        raise BizError(
            error_code="EXCEL_PARSE_FAILED",
            message="未识别到绩效得分列，请确保包含「绩效得分」「得分」或类似列",
        )

    db = SessionLocal()
    batch_id = f"batch_{uuid.uuid4().hex[:8]}"
    try:
        run = db.query(SalaryRun).filter(SalaryRun.id == salary_run_id).first()
        if not run:
            raise BizError(error_code="RESOURCE_NOT_FOUND", message="工资任务不存在")

        # 直接读取原始 Excel 数据
        from openpyxl import load_workbook
        wb = load_workbook(storage_path, data_only=True)
        ws = wb[sheet["sheet_name"]]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        data_start = header_row + 1
        matched_count = 0
        unmatched_names = []
        duplicate_count = 0
        error_records = []

        for idx in range(data_start, len(all_rows)):
            row = all_rows[idx]
            if not row or len(row) <= max(name_col, score_col):
                continue

            raw_name = row[name_col]
            if raw_name is None or str(raw_name).strip() == "":
                continue

            employee_name = str(raw_name).strip()

            raw_score = row[score_col]
            if raw_score is None:
                # 空得分 → 跳过（按 100% 处理）
                continue

            # 解析得分
            parsed = _parse_score(raw_score)
            if parsed is None:
                error_records.append({
                    "employee_name": employee_name,
                    "raw_value": str(raw_score),
                    "error": "无法解析的绩效得分",
                })
                continue

            score_value, parse_method = parsed

            # 范围校验
            if score_value < Decimal("0"):
                error_records.append({
                    "employee_name": employee_name,
                    "raw_value": str(raw_score),
                    "error": "绩效得分不能小于0",
                })
                continue
            if score_value > Decimal("1"):
                error_records.append({
                    "employee_name": employee_name,
                    "raw_value": str(raw_score),
                    "error": "绩效得分不能大于100%",
                })
                continue

            # 匹配员工
            emp = _match_employee(db, salary_run_id, employee_name)
            emp_id = emp.id if emp else None

            # 检查重复导入
            existing = db.query(PerformanceImportRecord).filter(
                PerformanceImportRecord.salary_run_id == salary_run_id,
                PerformanceImportRecord.employee_record_id == emp_id,
                PerformanceImportRecord.status != "DUPLICATE",
            ).first()
            if existing:
                # 覆盖之前的导入记录
                existing.score = score_value
                existing.parse_method = parse_method
                existing.status = "MATCHED"
                duplicate_count += 1
            else:
                # 新记录
                perf_rec = PerformanceImportRecord(
                    id=f"perf_{uuid.uuid4().hex[:8]}",
                    salary_run_id=salary_run_id,
                    employee_record_id=emp_id,
                    employee_name=employee_name,
                    score=score_value,
                    parse_method=parse_method,
                    batch_id=batch_id,
                    status="MATCHED" if emp_id else "UNMATCHED",
                    created_by=created_by,
                )
                db.add(perf_rec)

            if emp_id:
                # 更新员工记录的绩效得分
                emp.performance_score = score_value
                emp.performance_score_source = "IMPORTED"
                matched_count += 1
            else:
                unmatched_names.append(employee_name)

        db.commit()

        logger.info(
            "绩效导入完成: matched=%d, unmatched=%d, duplicate=%d, errors=%d",
            matched_count, len(unmatched_names), duplicate_count, len(error_records),
        )

        return {
            "batch_id": batch_id,
            "matched_count": matched_count,
            "unmatched_names": unmatched_names,
            "unmatched_count": len(unmatched_names),
            "duplicate_count": duplicate_count,
            "error_records": error_records,
            "error_count": len(error_records),
        }

    except BizError:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.error("绩效导入异常: %s", str(e)[:300], exc_info=True)
        raise BizError(
            error_code="EXCEL_PARSE_FAILED",
            message=f"绩效导入失败: {str(e)}",
        )
    finally:
        db.close()


def _detect_name_column(headers: list) -> int | None:
    """检测姓名/编号列"""
    name_keywords = ["姓名", "员工", "姓名", "员工编号", "工号", "名字", "员工姓名"]
    for i, h in enumerate(headers):
        if h:
            h_str = str(h).strip()
            for kw in name_keywords:
                if kw in h_str:
                    return i
    return None


def _detect_score_column(headers: list) -> int | None:
    """检测绩效得分列"""
    score_keywords = ["绩效", "得分", "评分", "系数", "绩效得分", "绩效分", "考核"]
    for i, h in enumerate(headers):
        if h:
            h_str = str(h).strip()
            for kw in score_keywords:
                if kw in h_str:
                    return i
    return None


def _parse_score(value) -> tuple[Decimal, str] | None:
    """
    解析绩效得分值。

    Returns:
        (Decimal 得分, str 解析方式) | None
        解析方式: PERCENT / DECIMAL / TEXT
    """
    if value is None:
        return None

    # 如果是 Decimal/float/int
    if isinstance(value, (Decimal, float, int)):
        val = Decimal(str(value))
        if val >= Decimal("1"):
            # 数值 90 → 90%
            return (val / Decimal("100"), "DECIMAL")
        else:
            # 数值 0.9 → 90%
            return (val, "DECIMAL")

    # 字符串
    text = str(value).strip().replace(" ", "")
    if not text:
        return None

    # "90%" 格式
    if "%" in text:
        try:
            num_text = text.replace("%", "").strip()
            num = Decimal(num_text)
            return (num / Decimal("100"), "PERCENT")
        except InvalidOperation:
            return None

    # "0.9" 或 "90" 格式
    try:
        num = Decimal(text)
        if num >= Decimal("1"):
            return (num / Decimal("100"), "TEXT")
        else:
            return (num, "TEXT")
    except InvalidOperation:
        return None


def _match_employee(
    db: SessionLocal,
    salary_run_id: str,
    employee_name: str,
) -> EmployeeRecord | None:
    """按姓名匹配员工记录"""
    from app.services.pool_service import clean_name
    clean = clean_name(employee_name)
    if not clean:
        return None

    emp = db.query(EmployeeRecord).filter(
        EmployeeRecord.salary_run_id == salary_run_id,
        EmployeeRecord.employee_name == clean,
        EmployeeRecord.status.in_(["NORMAL", "IGNORED"]),
    ).first()

    return emp
