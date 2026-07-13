"""
请假导入服务

负责：
- 解析企业微信请假 Excel
- 匹配员工（优先按员工编号，再按姓名）
- 汇总请假天数（同员工多条记录累加）
- 审批单号去重
- 更新员工月度请假天数
"""

import logging
import uuid
from datetime import datetime
from decimal import Decimal

from app.core.exceptions import BizError
from app.db.database import SessionLocal
from app.models import (
    EmployeeRecord, LeaveImportRecord, SalaryRun, CheckIssue,
)
from app.engines.importer.importer import parse_workbook, validate_file, save_file

logger = logging.getLogger(__name__)


def import_leave_records(
    file_bytes: bytes,
    original_name: str,
    salary_run_id: str,
    created_by: str,
) -> dict:
    """
    导入企业微信请假 Excel。

    Excel 至少包含：员工姓名 + 请假天数
    可能包含：员工编号、请假类型、请假日期、审批单号

    Returns:
        {
            "batch_id": str,
            "matched_count": int,
            "total_leave_days": str,
            "employee_count": int,
            "unmatched_names": [str],
            "duplicate_count": int,
            "error_records": [{...}],
        }
    """
    validate_file(file_bytes, original_name)
    storage_path, file_hash = save_file(file_bytes, original_name, salary_run_id)

    parse_result = parse_workbook(storage_path)
    if parse_result["sheet_count"] == 0:
        raise BizError(error_code="EXCEL_PARSE_FAILED", message="未找到工作表")

    sheet = parse_result["sheets"][0]
    headers = sheet["headers"]
    header_row = sheet["header_row_index"] or 0

    # 检测列
    name_col = _detect_column(headers, ["姓名", "员工", "员工姓名", "名字", "申请人"])
    emp_no_col = _detect_column(headers, ["员工编号", "工号", "编号", "员工ID"])
    days_col = _detect_column(headers, ["请假天数", "天数", "时长", "合计天数", "请假时长"])
    type_col = _detect_column(headers, ["请假类型", "类型", "假期类型"])
    start_col = _detect_column(headers, ["开始日期", "开始时间", "请假开始", "起始日期"])
    end_col = _detect_column(headers, ["结束日期", "结束时间", "请假结束", "截止日期"])
    approval_col = _detect_column(headers, ["审批单号", "审批编号", "单号", "申请单号", "审批号"])

    if name_col is None:
        raise BizError(
            error_code="EXCEL_PARSE_FAILED",
            message="未识别到员工姓名列，请确保包含「姓名」或「员工」列",
        )
    if days_col is None:
        raise BizError(
            error_code="EXCEL_PARSE_FAILED",
            message="未识别到请假天数列，请确保包含「请假天数」或「天数」列",
        )

    db = SessionLocal()
    batch_id = f"batch_{uuid.uuid4().hex[:8]}"
    try:
        run = db.query(SalaryRun).filter(SalaryRun.id == salary_run_id).first()
        if not run:
            raise BizError(error_code="RESOURCE_NOT_FOUND", message="工资任务不存在")

        from openpyxl import load_workbook
        wb = load_workbook(storage_path, data_only=True)
        ws = wb[sheet["sheet_name"]]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        data_start = header_row + 1
        matched_count = 0
        unmatched_names = []
        duplicate_count = 0
        error_records = []
        seen_approvals = set()
        seen_employee_days = {}  # {(emp_id, date): True} for dedup
        total_leave_days = Decimal("0")
        matched_employee_ids = set()

        for idx in range(data_start, len(all_rows)):
            row = all_rows[idx]
            if not row or len(row) <= max(name_col, days_col):
                continue

            raw_name = row[name_col]
            if raw_name is None or str(raw_name).strip() == "":
                continue
            employee_name = str(raw_name).strip()

            # 请假天数
            raw_days = row[days_col] if days_col < len(row) else None
            if raw_days is None:
                continue
            try:
                leave_days = Decimal(str(raw_days))
            except Exception:
                error_records.append({
                    "employee_name": employee_name,
                    "raw_value": str(raw_days),
                    "error": "请假天数无法解析",
                })
                continue

            if leave_days <= Decimal("0"):
                error_records.append({
                    "employee_name": employee_name,
                    "raw_value": str(raw_days),
                    "error": "请假天数必须大于0",
                })
                continue

            # 员工编号
            emp_no = str(row[emp_no_col]).strip() if emp_no_col is not None and emp_no_col < len(row) and row[emp_no_col] is not None else ""

            # 请假类型
            leave_type = str(row[type_col]).strip() if type_col is not None and type_col < len(row) and row[type_col] is not None else ""

            # 审批单号
            approval_id = ""
            if approval_col is not None and approval_col < len(row) and row[approval_col] is not None:
                approval_id = str(row[approval_col]).strip()

            # 日期信息
            leave_date = row[start_col] if start_col is not None and start_col < len(row) else None
            leave_start = leave_date
            leave_end = row[end_col] if end_col is not None and end_col < len(row) else None

            # 审批单号去重
            if approval_id:
                if approval_id in seen_approvals:
                    duplicate_count += 1
                    continue
                seen_approvals.add(approval_id)

            # 匹配员工 - 优先按编号，再按姓名
            emp = _match_employee(db, salary_run_id, emp_no, employee_name)
            emp_id = emp.id if emp else None

            # 同员工同日期去重
            if emp_id and leave_date:
                dedup_key = (emp_id, str(leave_date))
                if dedup_key in seen_employee_days:
                    duplicate_count += 1
                    continue
                seen_employee_days[dedup_key] = True

            # 保存导入记录
            leave_rec = LeaveImportRecord(
                id=f"leave_{uuid.uuid4().hex[:8]}",
                salary_run_id=salary_run_id,
                employee_record_id=emp_id,
                employee_name=employee_name,
                employee_no=emp_no or None,
                leave_date=leave_date if isinstance(leave_date, (datetime, type(None))) else None,
                leave_start=leave_start if isinstance(leave_start, (datetime, type(None))) else None,
                leave_end=leave_end if isinstance(leave_end, (datetime, type(None))) else None,
                leave_days=leave_days,
                leave_type=leave_type,
                approval_id=approval_id or None,
                batch_id=batch_id,
                status="MATCHED" if emp_id else "UNMATCHED",
                created_by=created_by,
            )
            db.add(leave_rec)

            if emp_id:
                matched_count += 1
                matched_employee_ids.add(emp_id)
                total_leave_days += leave_days
            else:
                unmatched_names.append(employee_name)

        # 汇总更新每位员工的请假总天数
        for emp_id in matched_employee_ids:
            from app.services.attendance_service import calculate_leave_summary
            emp_records = db.query(LeaveImportRecord).filter(
                LeaveImportRecord.salary_run_id == salary_run_id,
                LeaveImportRecord.employee_record_id == emp_id,
                LeaveImportRecord.status == "MATCHED",
            ).all()
            total = calculate_leave_summary(emp_records)

            emp = db.query(EmployeeRecord).filter(EmployeeRecord.id == emp_id).first()
            if emp:
                emp.leave_days = total

        db.commit()

        logger.info(
            "请假导入完成: matched=%d, employees=%d, total_days=%s, unmatched=%d, duplicates=%d, errors=%d",
            matched_count, len(matched_employee_ids), str(total_leave_days),
            len(unmatched_names), duplicate_count, len(error_records),
        )

        return {
            "batch_id": batch_id,
            "matched_count": matched_count,
            "employee_count": len(matched_employee_ids),
            "total_leave_days": str(total_leave_days),
            "unmatched_names": unmatched_names,
            "unmatched_count": len(unmatched_names),
            "duplicate_count": duplicate_count,
            "error_records": error_records,
            "error_count": len(error_records),
        }

    except BizError:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.error("请假导入异常: %s", str(e)[:300], exc_info=True)
        raise BizError(
            error_code="EXCEL_PARSE_FAILED",
            message=f"请假导入失败: {str(e)}",
        )
    finally:
        db.close()


def _detect_column(headers: list, keywords: list[str]) -> int | None:
    """检测列名"""
    for i, h in enumerate(headers):
        if h:
            h_str = str(h).strip()
            for kw in keywords:
                if kw in h_str:
                    return i
    return None


def _match_employee(
    db: SessionLocal,
    salary_run_id: str,
    employee_no: str,
    employee_name: str,
) -> EmployeeRecord | None:
    """按编号或姓名匹配员工"""
    from app.services.pool_service import clean_name
    clean_name_val = clean_name(employee_name)
    if not clean_name_val:
        return None

    # 优先按编号匹配
    if employee_no:
        emp = db.query(EmployeeRecord).filter(
            EmployeeRecord.salary_run_id == salary_run_id,
            EmployeeRecord.employee_no == employee_no,
        ).first()
        if emp:
            return emp

    # 按姓名匹配
    emp = db.query(EmployeeRecord).filter(
        EmployeeRecord.salary_run_id == salary_run_id,
        EmployeeRecord.employee_name == clean_name_val,
        EmployeeRecord.status.in_(["NORMAL", "IGNORED"]),
    ).first()

    return emp
