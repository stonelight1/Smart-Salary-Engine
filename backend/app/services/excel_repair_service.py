"""
Excel 文件修复服务

职责：
1. 验证 xlsx 文件结构完整性（ZIP 内部 XML）
2. 使用 LibreOffice Headless 自动修复损坏的工作簿
3. 使用 python-calamine 兼容降级读取（仅保留值）
4. 诊断损坏的内部 XML 文件（日志记录，不暴露给前端）
5. 管理临时文件生命周期
"""

import io
import os
import shutil
import logging
import subprocess
import tempfile
import time
import uuid
import zipfile
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# 最大解压后大小（200MB）
MAX_ARCHIVE_SIZE = 200 * 1024 * 1024
# LibreOffice 超时（秒）
LIBREOFFICE_TIMEOUT = 60
# LibreOffice 最大并发数
MAX_CONCURRENT_REPAIR = 4
_repair_semaphore = None


def _get_repair_semaphore():
    """获取修复信号量（延迟初始化）"""
    global _repair_semaphore
    if _repair_semaphore is None:
        import threading
        _repair_semaphore = threading.Semaphore(MAX_CONCURRENT_REPAIR)
    return _repair_semaphore


def is_libreoffice_available() -> bool:
    """检查 LibreOffice 是否可用"""
    return shutil.which("soffice") is not None


# ============================================================
#  公开接口
# ============================================================

ValidationResult = dict[str, Any]
RepairResult = dict[str, Any]


def validate_xlsx(file_bytes: bytes) -> ValidationResult:
    """
    验证 xlsx 文件 ZIP 内部结构完整性。

    Returns:
        {"valid": True}
        {"valid": False, "error": "ERROR_CODE", "message": "..."}
    """
    if not file_bytes or len(file_bytes) == 0:
        return {"valid": False, "error": "EMPTY_FILE", "message": "文件为空"}

    if file_bytes[:2] != b"PK":
        return {"valid": False, "error": "NOT_XLSX", "message": "文件不是有效的 ZIP/Excel 格式"}

    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as zf:
            names = zf.namelist()

            # 检查基本结构
            if "[Content_Types].xml" not in names:
                return {"valid": False, "error": "MISSING_CONTENT_TYPES", "message": "缺少 [Content_Types].xml"}

            if "xl/workbook.xml" not in names:
                return {"valid": False, "error": "MISSING_WORKBOOK", "message": "缺少 xl/workbook.xml"}

            # 检查 ZIP bomb / 过大
            total_size = 0
            for info in zf.infolist():
                if info.file_size > 0 and info.compress_size > 0:
                    ratio = info.file_size / info.compress_size
                    if ratio > 200:
                        return {"valid": False, "error": "ZIP_BOMB", "message": "文件压缩比异常，可能为 ZIP bomb"}
                total_size += info.file_size

            if total_size > MAX_ARCHIVE_SIZE:
                return {"valid": False, "error": "TOO_LARGE", "message": "解压后文件大小超过限制"}

            # 检查路径穿越
            for name in names:
                resolved = Path("/", name).resolve()
                if not str(resolved).startswith("/"):
                    return {"valid": False, "error": "PATH_TRAVERSAL", "message": "文件包含路径穿越条目"}

        return {"valid": True}

    except zipfile.BadZipFile:
        return {"valid": False, "error": "BAD_ZIP", "message": "文件不是有效的 ZIP 压缩包"}
    except Exception as e:
        logger.warning("xlsx 结构验证异常: %s", str(e)[:200])
        return {"valid": False, "error": "VALIDATION_ERROR", "message": "文件校验过程中发生未知错误"}


def classify_parse_error(error: Exception) -> str:
    """
    根据解析错误类型返回分类。
    返回: "INVALID_XML" | "WORKSHEET_ERROR" | "SHARED_STRINGS" | "OTHER"
    """
    msg = str(error).lower()

    if "invalid xml" in msg or "xml" in msg or "parsing" in msg.lower():
        return "INVALID_XML"
    if "worksheet" in msg or "sheet" in msg or "workbook" in msg:
        return "WORKSHEET_ERROR"
    if "sharedstrings" in msg:
        return "SHARED_STRINGS"
    if "badzipfile" in msg or "not a zip" in msg:
        return "BAD_ZIP"
    if "file is not a zip" in msg or "could not read" in msg:
        return "WORKBOOK_DAMAGED"

    return "OTHER"


def try_repair_libreoffice(
    source_path: str,
    temp_dir: str,
    timeout: int = LIBREOFFICE_TIMEOUT,
) -> RepairResult:
    """
    使用 LibreOffice Headless 自动修复并重新保存 xlsx。

    Args:
        source_path: 原始文件路径
        temp_dir: 临时工作目录
        timeout: 超时秒数

    Returns:
        {"success": True, "output_path": "..."}
        {"success": False, "error": "..."}
    """
    if not is_libreoffice_available():
        logger.warning("LibreOffice 不可用，跳过修复")
        return {"success": False, "error": "LIBREOFFICE_NOT_FOUND"}

    # 创建独立临时目录
    work_dir = Path(temp_dir) / f"lo_{uuid.uuid4().hex[:8]}"
    work_dir.mkdir(parents=True, exist_ok=True)

    # 复制源文件到工作目录（使用唯一名称避免覆盖）
    safe_filename = f"source_{uuid.uuid4().hex[:12]}.xlsx"
    source_copy = work_dir / safe_filename
    try:
        shutil.copy2(source_path, source_copy)
    except Exception as e:
        logger.error("复制源文件到临时目录失败: %s", str(e)[:200])
        _cleanup_dir(work_dir)
        return {"success": False, "error": "FILE_COPY_FAILED"}

    output_dir = work_dir / "output"
    output_dir.mkdir(parents=True, exist_ok=True)

    # 独立 LibreOffice profile 避免并发锁定
    profile_dir = work_dir / "profile"
    profile_dir.mkdir(parents=True, exist_ok=True)
    user_install = f"file://{profile_dir.resolve().as_posix()}"

    _acquired = _get_repair_semaphore().acquire(timeout=timeout)
    if not _acquired:
        _cleanup_dir(work_dir)
        return {"success": False, "error": "REPAIR_BUSY", "message": "修复服务繁忙，请稍后重试"}

    start_time = time.time()
    try:
        result = subprocess.run(
            [
                "soffice",
                "--headless",
                f"-env:UserInstallation={user_install}",
                "--convert-to",
                "xlsx:Calc MS Excel 2007 XML",
                "--outdir",
                str(output_dir.resolve()),
                str(source_copy.resolve()),
            ],
            timeout=timeout,
            capture_output=True,
            text=True,
            env={**os.environ, "HOME": str(profile_dir.resolve())},
        )

        elapsed = time.time() - start_time
        logger.info(
            "LibreOffice 修复耗时 %.2fs, returncode=%d, stderr=%.200s",
            elapsed, result.returncode, result.stderr or "",
        )

        if result.returncode != 0:
            logger.error("LibreOffice 修复失败: returncode=%d, stderr=%s", result.returncode, result.stderr)
            return {"success": False, "error": "LIBREOFFICE_FAILED"}

        # 查找输出文件（LibreOffice 输出文件名与源文件相同但路径在 outdir）
        output_files = list(output_dir.glob("*.xlsx"))
        if not output_files:
            logger.error("LibreOffice 未生成输出文件")
            return {"success": False, "error": "NO_OUTPUT"}

        output_path = str(output_files[0])

        # 验证输出文件
        if os.path.getsize(output_path) == 0:
            logger.error("LibreOffice 输出文件为空")
            return {"success": False, "error": "EMPTY_OUTPUT"}

        return {"success": True, "output_path": output_path}

    except subprocess.TimeoutExpired:
        logger.error("LibreOffice 修复超时（%ds）", timeout)
        return {"success": False, "error": "TIMEOUT"}
    except FileNotFoundError:
        logger.error("soffice 命令不存在")
        return {"success": False, "error": "LIBREOFFICE_NOT_FOUND"}
    except Exception as e:
        logger.error("LibreOffice 修复异常: %s", str(e)[:300])
        return {"success": False, "error": "UNEXPECTED"}
    finally:
        _get_repair_semaphore().release()
        _cleanup_dir(work_dir)


def try_rebuild_value_only(
    source_path: str,
    temp_dir: str,
) -> RepairResult:
    """
    使用 python-calamine 兼容模式读取单元格值，重建标准 xlsx。

    此方案仅保留单元格值，公式、样式、图片、批注等可能丢失。

    Returns:
        {"success": True, "output_path": "...", "warnings": [...]}
        {"success": False, "error": "..."}
    """
    try:
        from python_calamine import CalamineWorkbook
    except ImportError:
        logger.warning("python-calamine 未安装，跳过降级读取")
        return {"success": False, "error": "CALAMINE_NOT_INSTALLED"}

    start_time = time.time()
    output_path = os.path.join(temp_dir, f"rebuilt_{uuid.uuid4().hex[:12]}.xlsx")

    try:
        wb = CalamineWorkbook.from_path(source_path)
    except Exception as e:
        logger.error("python-calamine 打开文件失败: %s", str(e)[:200])
        return {"success": False, "error": "CALAMINE_OPEN_FAILED"}

    try:
        from openpyxl import Workbook as NewWorkbook
        new_wb = NewWorkbook()
        # 移除默认 Sheet
        default_ws = new_wb.active
        if default_ws:
            new_wb.remove(default_ws)

        sheet_names = wb.sheet_names
        for sname in sheet_names:
            try:
                sheet = wb.get_sheet_by_name(sname)
                rows_data = sheet.to_python()
            except Exception as e:
                logger.warning("calamine 读取 sheet[%s] 失败: %s", sname, str(e)[:200])
                continue

            ws = new_wb.create_sheet(title=sname)
            for row_idx, row in enumerate(rows_data, start=1):
                for col_idx, val in enumerate(row, start=1):
                    # Calamine returns None for empty cells
                    if val is not None:
                        ws.cell(row=row_idx, column=col_idx, value=val)

        new_wb.save(output_path)
        elapsed = time.time() - start_time

        if os.path.getsize(output_path) == 0:
            return {"success": False, "error": "EMPTY_OUTPUT"}

        logger.info("calamine 降级重建成功, sheets=%d, 耗时=%.2fs", len(sheet_names), elapsed)

        return {
            "success": True,
            "output_path": output_path,
            "warnings": ["部分公式、样式、图片或批注可能未保留，请确认导入数据无误后继续"],
            "method": "VALUE_ONLY_REBUILD",
        }

    except Exception as e:
        logger.error("calamine 重建 xlsx 失败: %s", str(e)[:300])
        return {"success": False, "error": "REBUILD_FAILED"}


def diagnose_workbook(file_path: str) -> dict[str, Any]:
    """
    诊断工作簿内部 XML 状态（仅写入日志，不返回前端）。

    Returns:
        {"parts_ok": [...], "parts_failed": [{"name": "...", "error": "..."}]}
    """
    parts_ok = []
    parts_failed = []

    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            xml_files = [n for n in zf.namelist() if n.endswith(".xml")]

            for xml_path in xml_files:
                try:
                    content = zf.read(xml_path)
                    # 尝试解析 XML
                    import xml.etree.ElementTree as ET
                    ET.fromstring(content)
                    parts_ok.append(xml_path)
                except Exception as e:
                    parts_failed.append({
                        "name": xml_path,
                        "error": str(e)[:200],
                    })

    except Exception as e:
        logger.error("工作簿诊断失败: %s", str(e)[:200])
        return {"parts_ok": [], "parts_failed": [{"name": "GENERAL", "error": str(e)[:200]}]}

    logger.info(
        "工作簿诊断完成: XML 正常=%d, 异常=%d",
        len(parts_ok), len(parts_failed),
    )
    for pf in parts_failed:
        logger.warning("损坏部件 [%s]: %s", pf["name"], pf["error"])

    return {"parts_ok": parts_ok, "parts_failed": parts_failed}


def cleanup_temp_files(temp_dir: str) -> None:
    """清理临时目录"""
    _cleanup_dir(Path(temp_dir))


def _cleanup_dir(dir_path: Path) -> None:
    """安全删除目录"""
    try:
        if dir_path.exists():
            shutil.rmtree(dir_path, ignore_errors=True)
    except Exception as e:
        logger.warning("清理临时目录失败 %s: %s", dir_path, str(e)[:200])
