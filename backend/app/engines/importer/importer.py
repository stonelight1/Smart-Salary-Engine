"""Excel 导入引擎

负责：
- 校验文件类型 (.xlsx)
- 保存原始文件到 uploads 目录
- 解析 workbook、Sheet、表头、数据预览
"""

import hashlib
import io
import logging
import os
import uuid
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from app.core.exceptions import BizError

logger = logging.getLogger(__name__)
IMPORT_BATCH_DIR = Path(__file__).resolve().parent.parent.parent.parent.parent / "uploads"


def _generate_file_id() -> str:
    return f"file_{uuid.uuid4().hex[:8]}"


def _generate_batch_id() -> str:
    return f"batch_{uuid.uuid4().hex[:8]}"


def validate_file(file_bytes: bytes, original_name: str) -> None:
    """校验文件类型和大小（增强版）"""
    # 1. 空文件检查
    if not file_bytes or len(file_bytes) == 0:
        raise BizError(
            error_code="EXCEL_EMPTY_FILE",
            message="上传文件为空",
        )

    # 2. 扩展名检查
    if not original_name.lower().endswith(".xlsx"):
        raise BizError(
            error_code="EXCEL_INVALID_EXTENSION",
            message="只支持 .xlsx 格式文件",
        )

    # 3. 大小检查
    if len(file_bytes) > 50 * 1024 * 1024:
        raise BizError(
            error_code="EXCEL_TOO_LARGE",
            message="文件大小超过 50MB 限制",
        )

    # 4. ZIP 魔术字节检查
    if file_bytes[:2] != b"PK":
        raise BizError(
            error_code="EXCEL_INVALID_FORMAT",
            message="文件格式不正确。该文件虽然使用了 .xlsx 后缀，但不是有效的 Excel 工作簿，请使用 Excel 重新另存为 .xlsx 后上传。",
        )

    # 5. ZIP 内部结构基础检查
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            names = zf.namelist()
            if "[Content_Types].xml" not in names:
                raise BizError(
                    error_code="EXCEL_INVALID_STRUCTURE",
                    message="Excel 文件内部结构不完整（缺少内容类型定义），请使用 Excel 重新保存后上传。",
                )
            if "xl/workbook.xml" not in names:
                raise BizError(
                    error_code="EXCEL_INVALID_STRUCTURE",
                    message="Excel 文件内部结构不完整（缺少工作簿定义），请使用 Excel 重新保存后上传。",
                )
    except zipfile.BadZipFile:
        raise BizError(
            error_code="EXCEL_CORRUPTED_ZIP",
            message="文件损坏或不是有效的 Excel 文件，请使用 Excel 打开并另存为 .xlsx 后重新上传。",
        )


def save_file(file_bytes: bytes, original_name: str, salary_run_id: str) -> tuple[str, str]:
    """保存原始文件到 uploads 目录，返回 (文件路径, 文件哈希)"""
    save_dir = IMPORT_BATCH_DIR / salary_run_id / "original"
    save_dir.mkdir(parents=True, exist_ok=True)

    file_hash = hashlib.sha256(file_bytes).hexdigest()[:16]
    file_id = _generate_file_id()
    ext = Path(original_name).suffix
    storage_name = f"{file_id}{ext}"
    storage_path = save_dir / storage_name

    with open(storage_path, "wb") as f:
        f.write(file_bytes)

    return str(storage_path), file_hash


def parse_workbook(file_path: str) -> dict[str, Any]:
    """解析 workbook，返回 Sheet 列表和预览数据"""
    try:
        wb = load_workbook(file_path, data_only=True, keep_links=False)
    except InvalidFileException as e:
        raise BizError(
            error_code="EXCEL_PARSE_FAILED",
            message=f"Excel 文件格式无法识别，请使用 Excel 重新另存为 .xlsx 后上传。",
        )
    except Exception as e:
        # 捕获 openpyxl 解析异常，判断是否为 XML 损坏类错误
        err_cls = _classify_openpyxl_error(e)
        if err_cls == "INVALID_XML":
            raise BizError(
                error_code="EXCEL_PARSE_FAILED",
                message=f"检测到 Excel 文件内部结构异常，正在尝试自动修复……",
            )
        raise BizError(
            error_code="EXCEL_PARSE_FAILED",
            message=f"Excel 文件解析失败。",
        )

    sheets = []
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=False))
        if not rows:
            sheets.append({
                "sheet_name": sheet_name,
                "row_count": 0,
                "header_row_index": None,
                "headers": [],
                "preview_rows": [],
            })
            continue

        # 扫描表头行
        header_row_index, headers = _detect_header_row(rows)
        total_rows += len(rows)

        # 取最多 5 行预览数据
        preview_start = header_row_index + 1 if header_row_index is not None else 1
        preview = _extract_row_values(rows[preview_start:preview_start + 5])

        sheets.append({
            "sheet_name": sheet_name,
            "row_count": len(rows),
            "header_row_index": header_row_index,
            "headers": headers,
            "preview_rows": preview,
        })

    wb.close()

    return {
        "sheet_count": len(sheets),
        "sheets": sheets,
        "total_rows": total_rows,
    }


def _classify_openpyxl_error(e: Exception) -> str:
    """分类 openpyxl 解析错误"""
    msg = str(e).lower()
    if "invalid xml" in msg or "xml" in msg or "parsing" in msg or "syntax" in msg:
        return "INVALID_XML"
    if "worksheet" in msg or "workbook" in msg or "sheet" in msg:
        return "WORKSHEET_ERROR"
    if "zip" in msg or "bad" in msg:
        return "ZIP_ERROR"
    return "OTHER"


def _detect_header_row(rows: list) -> tuple[int | None, list[str]]:
    """自动探测表头行（扫描前 10 行）"""
    for i, row in enumerate(rows[:10]):
        cells = [c for c in row if c.value is not None]
        non_empty = len(cells)

        if non_empty < 2:
            continue

        values = [str(c.value).strip() for c in cells]

        # 排除明显不是表头的行
        if any(kw in values for kw in ["制表人", "日期", "单位", "备注", "打印"]):
            continue

        # 检查是否包含已知字段关键词
        known_keywords = ["姓名", "员工", "工资", "基本", "部门", "金额", "出勤", "绩效", "扣款", "社保", "补贴"]
        keyword_hits = sum(1 for v in values if any(kw in v for kw in known_keywords))

        if keyword_hits >= 1:
            # 检查下一行是否像数据行（不是空行）
            if i + 1 < len(rows):
                next_cells = [c for c in rows[i + 1] if c.value is not None]
                if next_cells:
                    return i, [c.value for c in cells if c.value is not None]

    # fallback: 取第一个有 >=2 非空单元格的行
    for i, row in enumerate(rows[:5]):
        cells = [c for c in row if c.value is not None]
        if len(cells) >= 2:
            return i, [c.value for c in cells if c.value is not None]

    return None, []


def _extract_row_values(rows: list) -> list[list[str | None]]:
    """提取行中的值"""
    result = []
    for row in rows:
        values = []
        for cell in row:
            v = cell.value
            if v is not None:
                values.append(str(v))
            else:
                values.append(None)
        result.append(values)
    return result


def normalize_column_name(name: str) -> str:
    """列名规范化"""
    import re
    name = name.strip()
    name = name.replace("　", " ")  # 全角空格
    name = name.replace("\n", "").replace("\r", "")
    name = name.replace("（元）", "").replace("(元)", "")
    name = name.replace("（", "").replace("(", "")
    name = name.replace("）", "").replace(")", "")
    name = name.replace("：", "").replace(":", "")
    # 全角转半角
    result = []
    for c in name:
        code = ord(c)
        if 0xFF01 <= code <= 0xFF5E:
            result.append(chr(code - 0xFEE0))
        else:
            result.append(c)
    return "".join(result).strip()


def validate_file_is_xlsx(file_bytes: bytes) -> bool:
    return file_bytes[:2] == b"PK"
