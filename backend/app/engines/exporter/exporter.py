"""Excel 导出引擎

基于工资主表模板导出最终工资表、计算过程、异常报告、来源明细
"""

import datetime
import uuid
from pathlib import Path

from openpyxl import load_workbook

from app.core.exceptions import BizError
from app.db.database import SessionLocal
from app.models import (
    SalaryRun, ExportFile, WorkbookFile, EmployeeRecord,
    CalculationResult, CheckIssue, EmployeeFieldValue,
)

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent.parent
EXPORT_DIR = PROJECT_ROOT / "exports"


def export_salary(
    salary_run_id: str,
    calc_version: int,
    include_trace: bool = True,
    include_issues: bool = True,
    include_sources: bool = True,
    created_by: str = "admin",
) -> dict:
    """导出工资结果 Excel"""
    db = SessionLocal()
    try:
        run = db.query(SalaryRun).filter(SalaryRun.id == salary_run_id).first()
        if not run:
            raise BizError(error_code="RESOURCE_NOT_FOUND", message="任务不存在")

        if run.status not in ("CALCULATED", "CONFIRMED", "LOCKED", "EXPORTED"):
            raise BizError(
                error_code="RUN_STATUS_NOT_ALLOWED",
                message=f"当前状态不允许导出: {run.status}",
            )

        version = calc_version or run.current_calc_version

        # 获取工资主表模板
        main_file = db.query(WorkbookFile).join(
            SalaryRun, WorkbookFile.salary_run_id == SalaryRun.id
        ).filter(
            WorkbookFile.salary_run_id == salary_run_id,
        ).first()

        if not main_file:
            raise BizError(error_code="RESOURCE_NOT_FOUND", message="未找到工资主表模板文件")

        template_path = Path(main_file.storage_path)
        if not template_path.exists():
            raise BizError(error_code="EXCEL_PARSE_FAILED", message=f"模板文件不存在: {template_path}")

        # 复制模板
        save_dir = EXPORT_DIR / salary_run_id
        save_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"{run.payroll_month}_{run.name}_工资结果_v{version}_{timestamp}.xlsx"
        output_path = save_dir / file_name

        wb = load_workbook(template_path)
        ws = wb.active

        # 写入计算结果到主表
        employees = db.query(EmployeeRecord).filter(
            EmployeeRecord.salary_run_id == salary_run_id,
            EmployeeRecord.status == "NORMAL",
        ).all()

        # 确定写入位置: 在已有数据行后追加结果列
        # 先找到姓名列位置和最后一行
        max_row = ws.max_row
        name_col = None
        for col in range(1, ws.max_column + 1):
            cell_val = str(ws.cell(row=1, column=col).value or "")
            if "姓名" in cell_val or "员工" in cell_val:
                name_col = col
                break

        # 追加结果列标题
        result_headers = ["实发工资", "应发工资", "扣款合计"]
        result_cols = {}
        for i, header in enumerate(result_headers):
            col = ws.max_column + 1 + i
            ws.cell(row=1, column=col, value=header)
            result_cols[header] = col

        # 填入结果
        for emp in employees:
            results = db.query(CalculationResult).filter(
                CalculationResult.salary_run_id == salary_run_id,
                CalculationResult.employee_record_id == emp.id,
                CalculationResult.calc_version == version,
            ).all()

            result_map = {r.item_code: str(r.amount) for r in results}

            # 找到员工所在行
            row = None
            if name_col:
                for r in range(2, max_row + 1):
                    if str(ws.cell(row=r, column=name_col).value or "").strip() == emp.employee_name:
                        row = r
                        break

            if not row:
                max_row += 1
                row = max_row
                ws.cell(row=row, column=name_col or 1, value=emp.employee_name)

            ws.cell(row=row, column=result_cols["实发工资"], value=result_map.get("net_salary", "0.00"))
            ws.cell(row=row, column=result_cols["应发工资"], value=result_map.get("gross_salary", "0.00"))
            ws.cell(row=row, column=result_cols["扣款合计"], value=result_map.get("deduction_total", "0.00"))

        # 计算过程 Sheet
        if include_trace:
            trace_ws = wb.create_sheet("计算过程")
            trace_ws.append(["员工姓名", "工资项", "金额", "公式"])
            for emp in employees:
                results = db.query(CalculationResult).filter(
                    CalculationResult.salary_run_id == salary_run_id,
                    CalculationResult.employee_record_id == emp.id,
                    CalculationResult.calc_version == version,
                ).all()
                for r in results:
                    trace_ws.append([emp.employee_name, r.item_name, str(r.amount), r.formula])

        # 异常报告 Sheet
        if include_issues:
            issues_ws = wb.create_sheet("异常报告")
            issues_ws.append(["等级", "异常编码", "说明", "状态"])
            issues = db.query(CheckIssue).filter(
                CheckIssue.salary_run_id == salary_run_id,
            ).all()
            for iss in issues:
                issues_ws.append([iss.level, iss.issue_code, iss.message, iss.status])

        # 字段来源 Sheet
        if include_sources:
            sources_ws = wb.create_sheet("字段来源")
            sources_ws.append(["员工姓名", "字段", "值", "来源文件", "Sheet", "行", "列"])
            for emp in employees:
                fvs = db.query(EmployeeFieldValue).filter(
                    EmployeeFieldValue.employee_record_id == emp.id,
                ).all()
                for fv in fvs:
                    sources_ws.append([
                        emp.employee_name,
                        fv.field_code,
                        str(fv.value_decimal or fv.value_text or ""),
                        fv.source_file_id,
                        fv.source_sheet,
                        fv.source_row,
                        fv.source_column,
                    ])

        # 保存
        wb.save(str(output_path))
        wb.close()

        # 保存导出记录
        export_id = f"export_{uuid.uuid4().hex[:8]}"
        export_record = ExportFile(
            id=export_id,
            salary_run_id=salary_run_id,
            calc_version=version,
            template_file_id=main_file.id,
            file_name=file_name,
            storage_path=str(output_path),
            created_by=created_by,
        )
        db.add(export_record)

        run.status = "EXPORTED"
        db.commit()

        return {
            "export_file_id": export_id,
            "file_name": file_name,
            "download_url": f"/api/v1/export-files/{export_id}/download",
        }
    except BizError:
        raise
    except Exception as e:
        raise BizError(
            error_code="EXCEL_PARSE_FAILED",
            message=f"导出失败: {str(e)}",
        )
    finally:
        db.close()
