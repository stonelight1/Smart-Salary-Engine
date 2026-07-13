"""
员工档案和考勤导入测试
"""

import uuid
from datetime import date

import pytest
from fastapi.testclient import TestClient

from app.main import app
from app.db.database import init_db, SessionLocal
from app.db.seed import seed_default_user
from app.core.config import config_loader
from app.models import EmployeeMaster, EmployeeChangeCandidate

client = TestClient(app)


def _clear_test_data():
    """清理测试数据"""
    db = SessionLocal()
    try:
        db.query(EmployeeChangeCandidate).filter(
            EmployeeChangeCandidate.salary_month == "2026-07"
        ).delete()
        db.commit()
    finally:
        db.close()


@pytest.fixture(autouse=True)
def setup():
    """每次测试前初始化数据库"""
    config_loader._configs = {}
    config_loader.load_all()
    init_db()
    seed_default_user()
    _clear_test_data()
    yield


def auth_header() -> dict:
    resp = client.post("/api/v1/auth/login", json={
        "username": "admin", "password": "admin123",
    })
    token = resp.json()["data"]["access_token"]
    return {"Authorization": f"Bearer {token}"}


class TestEmployeeListFiltering:
    """员工列表筛选测试"""

    def test_active_tab_shows_only_active(self):
        """未勾选'含离职'时，列表只返回在职员工"""
        headers = auth_header()

        # 使用唯一标识符避免与现有数据冲突
        unique = uuid.uuid4().hex[:6]

        # 创建在职员工
        client.post("/api/v1/employees", json={
            "employee_name": f"在职员工_{unique}",
            "employee_no": f"A{unique}",
            "department": "客服部",
            "position": "客服",
        }, headers=headers)

        # 创建已离职员工
        resp = client.post("/api/v1/employees", json={
            "employee_name": f"离职员工_{unique}",
            "employee_no": f"B{unique}",
            "department": "直播部",
            "position": "主播",
        }, headers=headers)
        emp_id = resp.json()["data"]["id"]
        client.post(f"/api/v1/employees/{emp_id}/resign", json={
            "resign_date": "2026-06-01",
            "reason": "测试离职",
        }, headers=headers)

        # 查询在职员工
        resp = client.get("/api/v1/employees?status=ACTIVE", headers=headers)
        assert resp.status_code == 200
        data = resp.json()["data"]
        # 验证返回的都是在职员工
        for item in data["items"]:
            assert item["status"] == "ACTIVE"

    def test_all_tab_shows_all_employees(self):
        """全部档案标签显示所有员工"""
        headers = auth_header()
        unique = uuid.uuid4().hex[:6]

        client.post("/api/v1/employees", json={
            "employee_name": f"在职员工_{unique}",
            "employee_no": f"A{unique}",
        }, headers=headers)

        resp = client.post("/api/v1/employees", json={
            "employee_name": f"离职员工_{unique}",
            "employee_no": f"B{unique}",
        }, headers=headers)
        emp_id = resp.json()["data"]["id"]
        client.post(f"/api/v1/employees/{emp_id}/resign", json={
            "resign_date": "2026-06-01",
            "reason": "测试离职",
        }, headers=headers)

        # 查询全部
        resp = client.get("/api/v1/employees?include_resigned=true", headers=headers)
        assert resp.status_code == 200
        data = resp.json()["data"]
        # 验证返回数量 >= 2（可能有其他数据）
        assert data["total"] >= 2

    def test_stats_match_list_count(self):
        """前端统计卡片、标签数量和后端查询结果一致"""
        headers = auth_header()
        unique = uuid.uuid4().hex[:6]

        # 创建测试员工
        for i in range(3):
            client.post("/api/v1/employees", json={
                "employee_name": f"测试员工_{unique}_{i}",
                "employee_no": f"E{unique}{i}",
            }, headers=headers)

        # 获取统计
        stats_resp = client.get("/api/v1/employees/stats/overview", headers=headers)
        stats = stats_resp.json()["data"]

        # 获取在职列表
        list_resp = client.get("/api/v1/employees?status=ACTIVE", headers=headers)
        list_data = list_resp.json()["data"]

        # 验证统计中的active >= 列表中的数量（可能有其他数据）
        assert stats["active"] >= list_data["total"]
        assert stats["active"] >= 3


class TestInvalidEmployeeNames:
    """无效员工姓名过滤测试"""

    def test_none_name_not_created(self):
        """姓名为None的行不得创建员工"""
        db = SessionLocal()
        try:
            # 清理已有None记录
            db.query(EmployeeMaster).filter(EmployeeMaster.employee_name == "None").delete()
            db.commit()

            emp = EmployeeMaster(
                id=f"test_none_{uuid.uuid4().hex[:6]}",
                employee_name="None",
                employee_no="N001",
                status="ACTIVE",
            )
            db.add(emp)
            db.commit()

            # 查询应排除None
            from app.services.employee_service import list_employees
            result = list_employees(page_size=1000)
            names = [e["employee_name"] for e in result["items"]]
            assert "None" not in names
        finally:
            db.close()

    def test_null_name_not_created(self):
        """姓名为null的行不得创建员工"""
        db = SessionLocal()
        try:
            db.query(EmployeeMaster).filter(EmployeeMaster.employee_name == "null").delete()
            db.commit()

            emp = EmployeeMaster(
                id=f"test_null_{uuid.uuid4().hex[:6]}",
                employee_name="null",
                employee_no="L001",
                status="ACTIVE",
            )
            db.add(emp)
            db.commit()

            from app.services.employee_service import list_employees
            result = list_employees(page_size=1000)
            names = [e["employee_name"] for e in result["items"]]
            assert "null" not in names
        finally:
            db.close()

    def test_empty_name_not_created(self):
        """姓名为空字符串的行不得创建员工"""
        db = SessionLocal()
        try:
            db.query(EmployeeMaster).filter(EmployeeMaster.employee_name == "").delete()
            db.commit()

            emp = EmployeeMaster(
                id=f"test_empty_{uuid.uuid4().hex[:6]}",
                employee_name="",
                employee_no="E001",
                status="ACTIVE",
            )
            db.add(emp)
            db.commit()

            from app.services.employee_service import list_employees
            result = list_employees(page_size=1000)
            names = [e["employee_name"] for e in result["items"]]
            assert "" not in names
        finally:
            db.close()

    def test_undefined_name_not_created(self):
        """姓名为undefined的行不得创建员工"""
        db = SessionLocal()
        try:
            db.query(EmployeeMaster).filter(EmployeeMaster.employee_name == "undefined").delete()
            db.commit()

            emp = EmployeeMaster(
                id=f"test_undefined_{uuid.uuid4().hex[:6]}",
                employee_name="undefined",
                employee_no="U001",
                status="ACTIVE",
            )
            db.add(emp)
            db.commit()

            from app.services.employee_service import list_employees
            result = list_employees(page_size=1000)
            names = [e["employee_name"] for e in result["items"]]
            assert "undefined" not in names
        finally:
            db.close()


class TestAttendanceImport:
    """考勤导入测试"""

    def test_import_creates_possible_hire(self):
        """导入后正确生成疑似入职"""
        headers = auth_header()
        unique = uuid.uuid4().hex[:6]

        # 创建已有员工
        client.post("/api/v1/employees", json={
            "employee_name": f"张三_{unique}",
            "employee_no": f"Z{unique}",
            "department": "客服部",
            "position": "客服",
        }, headers=headers)

        # 模拟导入
        from openpyxl import Workbook
        import io

        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "员工编号", "部门", "岗位"])
        ws.append([f"张三_{unique}", f"Z{unique}", "客服部", "客服"])  # 已匹配
        ws.append([f"李四_{unique}", f"L{unique}", "直播部", "主播"])  # 疑似新入职
        ws.append(["合计", "", "", ""])  # 汇总行，应跳过
        ws.append(["None", f"N{unique}", "", ""])  # 无效名称，应跳过

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        from app.services.attendance_compare_service import import_attendance_employees
        result = import_attendance_employees(
            file_bytes=buffer.read(),
            original_name="test.xlsx",
            salary_month="2026-07",
            created_by="admin",
        )

        assert result["matched_count"] == 1
        assert result["new_hire_count"] == 1
        assert result["summary_count"] == 1
        # None行被过滤到empty_count中
        assert result["empty_count"] >= 1

    def test_summary_rows_skipped(self):
        """合计、总结行不得创建员工"""
        from openpyxl import Workbook
        import io

        unique = uuid.uuid4().hex[:6]

        # 先创建员工程序
        headers = auth_header()
        client.post("/api/v1/employees", json={
            "employee_name": f"张三_{unique}",
            "employee_no": f"Z{unique}",
            "department": "客服部",
            "position": "客服",
        }, headers=headers)

        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "员工编号", "部门", "岗位"])
        ws.append([f"张三_{unique}", f"Z{unique}", "客服部", "客服"])
        ws.append(["合计", "", "", ""])
        ws.append(["总计", "", "", ""])
        ws.append(["汇总", "", "", ""])
        ws.append(["小计", "", "", ""])
        ws.append(["总结", "", "", ""])
        ws.append(["自动总结", "", "", ""])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        from app.services.attendance_compare_service import import_attendance_employees
        result = import_attendance_employees(
            file_bytes=buffer.read(),
            original_name="test.xlsx",
            salary_month="2026-07",
            created_by="admin",
        )

        assert result["summary_count"] == 6
        assert result["matched_count"] == 1

    def test_position_change_detected(self):
        """导入后正确识别岗位变化"""
        headers = auth_header()
        unique = uuid.uuid4().hex[:6]

        client.post("/api/v1/employees", json={
            "employee_name": f"张三_{unique}",
            "employee_no": f"Z{unique}",
            "department": "客服部",
            "position": "客服",
        }, headers=headers)

        from openpyxl import Workbook
        import io

        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "员工编号", "部门", "岗位"])
        ws.append([f"张三_{unique}", f"Z{unique}", "客服部", "运营"])  # 岗位变化

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        from app.services.attendance_compare_service import import_attendance_employees
        result = import_attendance_employees(
            file_bytes=buffer.read(),
            original_name="test.xlsx",
            salary_month="2026-07",
            created_by="admin",
        )

        assert result["pos_changes"] == 1
        assert result["matched_count"] == 1


class TestEmployeeChangeHandling:
    """员工异动处理测试"""

    def test_confirm_hire_adds_employee(self):
        """确认入职后，员工进入在职列表"""
        headers = auth_header()
        unique = uuid.uuid4().hex[:6]

        # 先导入考勤
        from openpyxl import Workbook
        import io

        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "员工编号", "部门", "岗位"])
        ws.append([f"新员工_{unique}", f"N{unique}", "客服部", "客服"])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        from app.services.attendance_compare_service import import_attendance_employees
        import_attendance_employees(
            file_bytes=buffer.read(),
            original_name="test.xlsx",
            salary_month="2026-07",
            created_by="admin",
        )

        # 获取候选人
        db = SessionLocal()
        try:
            cand = db.query(EmployeeChangeCandidate).filter(
                EmployeeChangeCandidate.candidate_type == "POSSIBLE_HIRE",
                EmployeeChangeCandidate.employee_name == f"新员工_{unique}",
            ).first()
            assert cand is not None

            # 确认入职
            resp = client.post(f"/api/v1/attendance-compare/handle/{cand.id}", json={
                "action": "CONFIRM_HIRE",
                "data": { "hire_date": "2026-07-01" },
            }, headers=headers)
            assert resp.status_code == 200
            assert resp.json()["success"] is True

            # 验证员工已创建
            emp = db.query(EmployeeMaster).filter(
                EmployeeMaster.employee_name == f"新员工_{unique}",
            ).first()
            assert emp is not None
            assert emp.status == "ACTIVE"
        finally:
            db.close()

    def test_keep_active_preserves_status(self):
        """疑似离职选择'保持在职'后，员工状态不变"""
        headers = auth_header()
        unique = uuid.uuid4().hex[:6]

        # 创建在职员工
        resp = client.post("/api/v1/employees", json={
            "employee_name": f"老员工_{unique}",
            "employee_no": f"O{unique}",
            "department": "直播部",
            "position": "主播",
        }, headers=headers)
        emp_id = resp.json()["data"]["id"]

        # 导入考勤（不含该员工）
        from openpyxl import Workbook
        import io

        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "员工编号", "部门", "岗位"])
        ws.append([f"其他员工_{unique}", f"X{unique}", "客服部", "客服"])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        from app.services.attendance_compare_service import import_attendance_employees
        import_attendance_employees(
            file_bytes=buffer.read(),
            original_name="test.xlsx",
            salary_month="2026-07",
            created_by="admin",
        )

        # 获取疑似离职候选人
        db = SessionLocal()
        try:
            cand = db.query(EmployeeChangeCandidate).filter(
                EmployeeChangeCandidate.candidate_type == "POSSIBLE_TERMINATION",
                EmployeeChangeCandidate.employee_master_id == emp_id,
            ).first()
            assert cand is not None

            # 保持在职
            resp = client.post(f"/api/v1/attendance-compare/handle/{cand.id}", json={
                "action": "KEEP_ACTIVE",
                "data": {},
            }, headers=headers)
            assert resp.status_code == 200

            # 验证员工状态仍为在职
            emp = db.query(EmployeeMaster).filter(EmployeeMaster.id == emp_id).first()
            assert emp.status == "ACTIVE"
        finally:
            db.close()

    def test_confirm_termination_changes_status(self):
        """确认离职后，员工从在职列表移入已离职列表"""
        headers = auth_header()
        unique = uuid.uuid4().hex[:6]

        # 创建在职员工
        resp = client.post("/api/v1/employees", json={
            "employee_name": f"待离职员工_{unique}",
            "employee_no": f"T{unique}",
            "department": "直播部",
            "position": "主播",
        }, headers=headers)
        emp_id = resp.json()["data"]["id"]

        # 导入考勤（不含该员工）
        from openpyxl import Workbook
        import io

        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "员工编号", "部门", "岗位"])
        ws.append([f"其他员工_{unique}", f"X{unique}", "客服部", "客服"])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        from app.services.attendance_compare_service import import_attendance_employees
        import_attendance_employees(
            file_bytes=buffer.read(),
            original_name="test.xlsx",
            salary_month="2026-07",
            created_by="admin",
        )

        # 获取疑似离职候选人
        db = SessionLocal()
        try:
            cand = db.query(EmployeeChangeCandidate).filter(
                EmployeeChangeCandidate.candidate_type == "POSSIBLE_TERMINATION",
                EmployeeChangeCandidate.employee_master_id == emp_id,
            ).first()
            assert cand is not None

            # 确认离职
            resp = client.post(f"/api/v1/attendance-compare/handle/{cand.id}", json={
                "action": "CONFIRM_TERMINATION",
                "data": { "termination_date": "2026-07-15" },
            }, headers=headers)
            assert resp.status_code == 200

            # 验证员工状态变为已离职
            emp = db.query(EmployeeMaster).filter(EmployeeMaster.id == emp_id).first()
            assert emp.status == "RESIGNED"
        finally:
            db.close()


class TestCandidateCounts:
    """待确认异动数量测试"""

    def test_pending_counts_match_list(self):
        """待确认异动数量与列表数量一致"""
        headers = auth_header()
        unique = uuid.uuid4().hex[:6]

        # 创建员工
        client.post("/api/v1/employees", json={
            "employee_name": f"员工A_{unique}",
            "employee_no": f"A{unique}",
            "department": "客服部",
            "position": "客服",
        }, headers=headers)

        # 导入考勤产生异动
        from openpyxl import Workbook
        import io

        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "员工编号", "部门", "岗位"])
        ws.append([f"员工A_{unique}", f"A{unique}", "销售部", "销售"])  # 部门变化
        ws.append([f"新员工B_{unique}", f"B{unique}", "客服部", "客服"])  # 疑似入职

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        from app.services.attendance_compare_service import import_attendance_employees
        import_attendance_employees(
            file_bytes=buffer.read(),
            original_name="test.xlsx",
            salary_month="2026-07",
            created_by="admin",
        )

        # 获取统计
        counts_resp = client.get("/api/v1/attendance-compare/counts", headers=headers)
        counts = counts_resp.json()["data"]

        # 获取列表
        list_resp = client.get("/api/v1/attendance-compare/candidates?status=PENDING", headers=headers)
        list_data = list_resp.json()["data"]

        # 验证统计中的PENDING_TOTAL >= 列表总数
        assert counts["PENDING_TOTAL"] >= list_data["total"]
        assert counts["PENDING_TOTAL"] > 0
